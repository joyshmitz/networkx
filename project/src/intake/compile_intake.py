#!/usr/bin/env python3
"""Compile per-person intake Excel sheets into questionnaire.yaml.

Usage:
    python -m network_methodology_sandbox.intake.compile_intake \
        project/examples/sample_object_01 --date 2026-04-02
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from network_methodology_sandbox.intake.workspace_manifest import refresh_workspace_manifest
from network_methodology_sandbox.intake.workspace_validation import (
    IntakeCommandError,
    ensure_compile_inputs,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PHASE_MAP: dict[str, int] = {
    "metadata": 1,
    "object_profile": 1,
    "critical_services": 1,
    "external_transport": 2,
    "security_access": 2,
    "time_sync": 2,
    "power_environment": 2,
    "resilience": 2,
    "operations": 3,
    "acceptance_criteria": 3,
    "governance": 3,
}

SERVICE_FIELDS: dict[str, str] = {
    "telemetry_required": "telemetry",
    "control_required": "control",
    "video_required": "video",
    "iiot_required": "iiot_edge",
    "local_archiving_required": "local_archiving",
}


def _parse_cli_date(raw: str) -> date:
    try:
        return date.fromisoformat(raw)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(
            f"Expected YYYY-MM-DD date, got: {raw!r}"
        ) from exc


# ---------------------------------------------------------------------------
# YAML helpers
# ---------------------------------------------------------------------------

def _load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    if not isinstance(data, dict):
        raise TypeError(f"Expected mapping in {path}, got {type(data).__name__}")
    return data


def _write_yaml(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        yaml.safe_dump(payload, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------

def _parse_value(raw: Any) -> str | None:
    """Parse dropdown value: 'code \u2014 label' \u2192 'code'.  Else return as-is."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if " \u2014 " in text:
        return text.split(" \u2014 ")[0].strip()
    return text


def _derive_status(
    value: str | None, explicit_status: str | None,
) -> tuple[str, bool]:
    """Auto-derive status.  Returns (status, has_warning).

    Warning is True when value is filled but user also set status=tbd.
    """
    has_value = value is not None and str(value).strip() != ""
    status = (explicit_status or "").strip().lower()

    if has_value and not status:
        return "answered", False
    if has_value and status == "tbd":
        return "answered", True  # suspicious: value present but marked tbd
    if has_value and status == "not_applicable":
        return "answered", False  # value takes precedence
    if not has_value and status == "tbd":
        return "tbd", False
    if not has_value and status == "not_applicable":
        return "not_applicable", False
    if not has_value and not status:
        return "unanswered", False
    return "answered" if has_value else "unanswered", False


def parse_xlsx(xlsx_path: Path) -> tuple[str, dict[str, dict], list[str]]:
    """Parse an intake xlsx.

    Returns (person_id, fields_dict, warnings).
    fields_dict: {field_id: {value, raw_value, status, comment, source_ref}}
    """
    person_id = xlsx_path.stem
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["intake"]

    fields: dict[str, dict] = {}
    warnings: list[str] = []

    for row in range(7, ws.max_row + 1):
        fid_cell = ws.cell(row, 1).value
        if fid_cell is None:
            continue
        fid = str(fid_cell).strip()
        if not fid:
            continue

        raw_value = ws.cell(row, 5).value
        raw_status = ws.cell(row, 6).value
        comment = ws.cell(row, 7).value
        source_ref = ws.cell(row, 8).value

        parsed = _parse_value(raw_value)
        explicit = str(raw_status).strip() if raw_status else None
        status, warned = _derive_status(parsed, explicit)

        if warned:
            warnings.append(
                f"{person_id}: field '{fid}' has value '{parsed}' "
                f"but status is 'tbd'"
            )

        fields[fid] = {
            "value": parsed,
            "raw_value": str(raw_value) if raw_value is not None else None,
            "status": status,
            "comment": str(comment).strip() if comment else None,
            "source_ref": str(source_ref).strip() if source_ref else None,
        }

    return person_id, fields, warnings


# ---------------------------------------------------------------------------
# Validators
# ---------------------------------------------------------------------------

def _validate_values(
    all_fields: dict[str, dict],
    field_index: dict[str, dict],
    value_dicts: dict[str, list[dict]],
) -> list[str]:
    """Validate parsed values against allowed vocabularies."""
    errors: list[str] = []
    for fid, entry in all_fields.items():
        if entry["status"] != "answered":
            continue
        value = entry["value"]
        if value is None:
            continue

        field_def = field_index.get(fid)
        if not field_def:
            continue
        field_type = field_def.get("type", "enum")

        if field_type == "string":
            continue

        if field_type == "integer":
            try:
                int(value)
            except (ValueError, TypeError):
                errors.append(f"Field '{fid}': expected integer, got '{value}'")
            continue

        # enum
        dict_ref = field_def.get("allowed_values_ref")
        if not dict_ref:
            continue
        entries = value_dicts.get(dict_ref, [])
        allowed = {e["value_code"] for e in entries} | {"tbd"}
        if value not in allowed:
            errors.append(
                f"Field '{fid}': value '{value}' not in {dict_ref} "
                f"(allowed: {sorted(allowed)})"
            )

    return errors


def _check_conflicts(
    person_fields: dict[str, dict[str, dict]],
) -> list[str]:
    """Detect field answered by more than one person."""
    field_owners: dict[str, list[str]] = {}
    for pid, fields in person_fields.items():
        for fid, entry in fields.items():
            if entry["status"] == "answered":
                field_owners.setdefault(fid, []).append(pid)

    return [
        f"Field '{fid}' answered by multiple persons: {owners}"
        for fid, owners in field_owners.items()
        if len(owners) > 1
    ]


# ---------------------------------------------------------------------------
# Questionnaire builder
# ---------------------------------------------------------------------------

def _build_questionnaire(
    all_fields: dict[str, dict],
    field_index: dict[str, dict],
    core_data: dict,
) -> dict[str, Any]:
    """Reconstruct questionnaire.yaml in section order."""
    q: dict[str, Any] = {"version": "0.2.0"}

    for section in core_data["sections"]:
        sid = section["id"]
        if sid == "known_unknowns":
            continue

        section_data: dict[str, Any] = {}
        for fid in section.get("fields", []):
            entry = all_fields.get(fid)
            if not entry or entry["status"] == "unanswered":
                continue

            status = entry["status"]
            value = entry["value"]

            if status == "answered":
                field_def = field_index.get(fid, {})
                if field_def.get("type") == "integer":
                    try:
                        section_data[fid] = int(value)
                    except (ValueError, TypeError):
                        section_data[fid] = value
                else:
                    section_data[fid] = value
            elif status in ("tbd", "not_applicable"):
                section_data[fid] = "tbd"

        if section_data:
            q[sid] = section_data

    q["known_unknowns"] = {}
    return q


# ---------------------------------------------------------------------------
# Response YAML writer
# ---------------------------------------------------------------------------

def _write_response_yaml(
    path: Path,
    person_id: str,
    fields: dict[str, dict],
    compiled_on: date,
) -> None:
    payload: dict[str, Any] = {
        "person_id": person_id,
        "compiled_at": compiled_on.isoformat(),
        "fields": {},
    }
    for fid in sorted(fields):
        entry = fields[fid]
        rec: dict[str, Any] = {
            "value": entry["value"],
            "status": entry["status"],
        }
        if entry.get("comment"):
            rec["comment"] = entry["comment"]
        if entry.get("source_ref"):
            rec["source_ref"] = entry["source_ref"]
        payload["fields"][fid] = rec

    _write_yaml(path, payload)


# ---------------------------------------------------------------------------
# Intake status writers
# ---------------------------------------------------------------------------

def _count_statuses(fields: dict[str, dict]) -> dict[str, int]:
    counts = {"answered": 0, "tbd": 0, "unanswered": 0, "not_applicable": 0}
    for entry in fields.values():
        s = entry["status"]
        if s in counts:
            counts[s] += 1
    return counts


def _build_intake_status_yaml(
    object_id: str,
    all_fields: dict[str, dict],
    person_fields: dict[str, dict[str, dict]],
    role_data: dict,
    field_to_section: dict[str, str],
    compiled_on: date,
) -> dict[str, Any]:
    totals = _count_statuses(all_fields)
    totals["total"] = len(all_fields)

    # person roles lookup
    person_roles: dict[str, list[str]] = {}
    for a in role_data.get("assignments", []):
        pid = a["person_id"]
        person_roles.setdefault(pid, [])
        for r in a["roles"]:
            if r not in person_roles[pid]:
                person_roles[pid].append(r)

    per_person: dict[str, dict] = {}
    for pid, fields in sorted(person_fields.items()):
        c = _count_statuses(fields)
        per_person[pid] = {
            "roles": sorted(person_roles.get(pid, [])),
            "owned": len(fields),
            **c,
        }

    fields_detail: dict[str, dict] = {}
    for fid, entry in sorted(
        all_fields.items(),
        key=lambda x: (field_to_section.get(x[0], ""), x[0]),
    ):
        fields_detail[fid] = {
            "section": field_to_section.get(fid, ""),
            "person_id": entry.get("person_id", ""),
            "status": entry["status"],
            "value": entry["value"],
        }

    return {
        "object_id": object_id,
        "compiled_at": compiled_on.isoformat(),
        "totals": totals,
        "per_person": per_person,
        "fields": fields_detail,
    }


def _phase_readiness(
    all_fields: dict[str, dict], field_to_section: dict[str, str],
) -> dict[int, dict[str, int]]:
    phases: dict[int, dict[str, int]] = {
        p: {"answered": 0, "tbd": 0, "unanswered": 0, "not_applicable": 0}
        for p in (1, 2, 3)
    }
    for fid, entry in all_fields.items():
        section = field_to_section.get(fid, "")
        phase = PHASE_MAP.get(section, 0)
        if phase in phases:
            s = entry["status"]
            if s in phases[phase]:
                phases[phase][s] += 1
    return phases


def _scope_summary(all_fields: dict[str, dict]) -> dict[str, Any]:
    summary: dict[str, Any] = {}
    for fid, key in [("object_type", "object_type"), ("criticality_class", "criticality")]:
        entry = all_fields.get(fid)
        if entry and entry["status"] == "answered":
            summary[key] = entry["value"]
    services = [
        svc for fid, svc in SERVICE_FIELDS.items()
        if all_fields.get(fid, {}).get("status") == "answered"
        and all_fields[fid]["value"] == "yes"
    ]
    if services:
        summary["services"] = services
    return summary


def _write_intake_status_md(
    path: Path,
    object_id: str,
    all_fields: dict[str, dict],
    person_fields: dict[str, dict[str, dict]],
    role_data: dict,
    field_index: dict[str, dict],
    field_to_section: dict[str, str],
    compiled_on: date,
) -> None:
    total = len(all_fields)
    counts = _count_statuses(all_fields)
    pct = counts["answered"] * 100 // total if total else 0

    # person roles
    person_roles: dict[str, set[str]] = {}
    for a in role_data.get("assignments", []):
        pid = a["person_id"]
        person_roles.setdefault(pid, set()).update(a["roles"])

    scope = _scope_summary(all_fields)
    phases = _phase_readiness(all_fields, field_to_section)
    phase_names = {1: "Identity", 2: "Constraints", 3: "Operations"}

    # Collect unassigned field ids (person_id == "_unassigned" or None)
    unassigned_fids = [
        fid for fid, e in all_fields.items()
        if e.get("person_id") in (None, "_unassigned")
    ]

    lines = [
        f"# Intake Status \u2014 {object_id}",
        "",
        f"Compiled at: {compiled_on.isoformat()}",
        "",
        f"Answered: {counts['answered']}/{total} ({pct}%) "
        f"| TBD: {counts['tbd']} | Unanswered: {counts['unanswered']} "
        f"| N/A: {counts['not_applicable']}",
        "",
    ]

    # Scope Summary
    if scope:
        lines.append("## Scope Summary")
        for k, v in scope.items():
            if isinstance(v, list):
                lines.append(f"- {k.replace('_', ' ').capitalize()}: {', '.join(v)}")
            else:
                lines.append(f"- {k.replace('_', ' ').capitalize()}: {v}")
        lines.append("")

    # Per Person
    lines.append("## Per Person")
    lines.append("| Person | Roles | Owned | Answered | TBD | Unanswered |")
    lines.append("|--------|-------|-------|----------|-----|------------|")
    for pid in sorted(person_fields):
        if pid.startswith("_"):
            continue
        fields = person_fields[pid]
        roles_str = ", ".join(sorted(person_roles.get(pid, set())))
        c = _count_statuses(fields)
        lines.append(
            f"| {pid} | {roles_str} | {len(fields)} "
            f"| {c['answered']} | {c['tbd']} | {c['unanswered']} |"
        )
    lines.append("")

    # Phase Readiness
    lines.append("## Phase Readiness")
    for p in (1, 2, 3):
        stats = phases[p]
        if stats["unanswered"] > 0:
            label = f"incomplete \u2014 {stats['unanswered']} unanswered"
        elif stats["tbd"] > 0:
            label = f"partial \u2014 {stats['tbd']} tbd"
        else:
            label = "complete"
        lines.append(f"- Phase {p} ({phase_names[p]}): {label}")
    lines.append("")

    # Unassigned Fields
    if unassigned_fids:
        lines.append("## Unassigned Fields")
        for fid in unassigned_fids:
            owner = field_index.get(fid, {}).get("owner_role", "unknown")
            lines.append(f"- {fid} (owner_role: {owner})")
        lines.append("")

    # Field Ownership Table
    lines.append("## Field Ownership Table")
    lines.append("| Field | Section | Owner Person | Status | Value |")
    lines.append("|-------|---------|-------------|--------|-------|")
    for fid, entry in sorted(
        all_fields.items(),
        key=lambda x: (field_to_section.get(x[0], ""), x[0]),
    ):
        section = field_to_section.get(fid, "")
        pid = entry.get("person_id") or ""
        val = entry.get("value") or ""
        lines.append(
            f"| {fid} | {section} | {pid} | {entry['status']} | {val} |"
        )
    lines.append("")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def compile_intake(
    workspace_path: Path,
    project_root: Path | None = None,
    compiled_on: date | None = None,
) -> dict[str, Any]:
    """Compile intake xlsx files into questionnaire.yaml + status reports.

    Returns a summary dict with keys: object_id, questionnaire, all_fields,
    person_fields, and warnings.
    Raises IntakeCommandError if user-facing validation errors are found.
    """
    if project_root is None:
        project_root = Path(__file__).resolve().parents[2]
    if compiled_on is None:
        compiled_on = date.today()
    workspace_path, xlsx_files = ensure_compile_inputs(workspace_path)

    # --- 1. Load specs ---
    fields_data = _load_yaml(
        project_root / "specs" / "dictionary" / "questionnaire_v2_fields.yaml"
    )
    values_data = _load_yaml(
        project_root / "specs" / "dictionary" / "questionnaire_v2_values.yaml"
    )
    core_data = _load_yaml(
        project_root / "specs" / "questionnaire" / "core_questionnaire_v2.yaml"
    )
    role_data = _load_yaml(workspace_path / "role_assignments.yaml")

    field_index = {f["field_id"]: f for f in fields_data["fields"]}
    value_dicts: dict[str, list[dict]] = values_data.get("dictionaries", {})

    field_to_section: dict[str, str] = {}
    for section in core_data["sections"]:
        for fid in section.get("fields", []):
            field_to_section[fid] = section["id"]

    object_id = role_data.get("object_id", workspace_path.name)

    # --- 2. Find and parse xlsx ---
    responses_dir = workspace_path / "intake" / "responses"
    person_fields: dict[str, dict[str, dict]] = {}
    all_warnings: list[str] = []

    for xlsx in xlsx_files:
        pid, fields, warnings = parse_xlsx(xlsx)
        person_fields[pid] = fields
        all_warnings.extend(warnings)

    # --- 3. Build global field mapping ---
    all_fields: dict[str, dict] = {}

    # Seed all spec fields as unanswered
    for fid in field_index:
        all_fields[fid] = {
            "value": None,
            "status": "unanswered",
            "comment": None,
            "source_ref": None,
            "person_id": None,
        }

    # Overlay with parsed responses
    for pid, fields in person_fields.items():
        for fid, entry in fields.items():
            all_fields[fid] = {**entry, "person_id": pid}

    # --- 4. Validate ---
    conflicts = _check_conflicts(person_fields)
    value_errors = _validate_values(all_fields, field_index, value_dicts)
    all_errors = conflicts + value_errors

    if all_errors:
        raise IntakeCommandError(
            "Compile validation failed:\n" + "\n".join(f"  - {e}" for e in all_errors)
        )

    # --- 5. Build questionnaire ---
    questionnaire = _build_questionnaire(all_fields, field_index, core_data)

    # --- 6. Write outputs ---
    # questionnaire.yaml
    _write_yaml(workspace_path / "questionnaire.yaml", questionnaire)

    # per-person .response.yaml
    for pid, fields in person_fields.items():
        resp_path = responses_dir / f"{pid}.response.yaml"
        _write_response_yaml(resp_path, pid, fields, compiled_on)

    # intake_status.yaml
    status_yaml = _build_intake_status_yaml(
        object_id, all_fields, person_fields, role_data, field_to_section,
        compiled_on,
    )
    _write_yaml(workspace_path / "reports" / "intake_status.yaml", status_yaml)

    # intake_status.md
    _write_intake_status_md(
        workspace_path / "reports" / "intake_status.md",
        object_id, all_fields, person_fields,
        role_data, field_index, field_to_section, compiled_on,
    )

    refresh_workspace_manifest(
        workspace_path,
        object_id=object_id,
        date_used=compiled_on.isoformat(),
        artifacts=[
            {
                "producer": "compile",
                "artifact_type": "questionnaire",
                "format": "yaml",
                "path": workspace_path / "questionnaire.yaml",
            },
            {
                "producer": "compile",
                "artifact_type": "intake_status",
                "format": "yaml",
                "path": workspace_path / "reports" / "intake_status.yaml",
            },
            {
                "producer": "compile",
                "artifact_type": "intake_status",
                "format": "markdown",
                "path": workspace_path / "reports" / "intake_status.md",
            },
        ],
    )

    # --- 7. Print warnings ---
    for w in all_warnings:
        print(f"WARNING: {w}", file=sys.stderr)

    return {
        "object_id": object_id,
        "questionnaire": questionnaire,
        "all_fields": all_fields,
        "person_fields": person_fields,
        "warnings": all_warnings,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compile intake sheets into questionnaire and status reports.",
    )
    parser.add_argument("workspace_path", help="Workspace directory to compile from.")
    parser.add_argument(
        "--date",
        dest="compiled_on",
        type=_parse_cli_date,
        help="Fixed compile date in YYYY-MM-DD format.",
    )
    args = parser.parse_args()

    try:
        result = compile_intake(Path(args.workspace_path), compiled_on=args.compiled_on)
    except IntakeCommandError as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)

    counts = _count_statuses(result["all_fields"])
    total = len(result["all_fields"])
    print(f"Compiled {result['object_id']}:")
    print(f"  {counts['answered']}/{total} answered, "
          f"{counts['tbd']} tbd, {counts['unanswered']} unanswered")
    print(f"  \u2192 questionnaire.yaml")
    print(f"  \u2192 reports/intake_status.yaml")
    print(f"  \u2192 reports/intake_status.md")


if __name__ == "__main__":
    main()
