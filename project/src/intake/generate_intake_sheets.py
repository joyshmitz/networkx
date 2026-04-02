#!/usr/bin/env python3
"""Generate per-person intake Excel sheets and guide.md from specs.

Usage:
    python src/intake/generate_intake_sheets.py examples/sample_object_01/
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Protection, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PHASE_MAP: dict[str, int] = {
    "metadata": 1,
    "object_profile": 1,
    "critical_services": 1,
    "external_transport": 2,
    "security_access": 2,
    "time_sync": 2,
    "power_environment": 2,
    "resilience": 2,
    "operations": 3,
    "acceptance_criteria": 3,
    "governance": 3,
}

STRICTNESS_FILLS: dict[str, PatternFill] = {
    "S4": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "S3": PatternFill(start_color="FFE0CC", end_color="FFE0CC", fill_type="solid"),
    "S2": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
    "S1": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
}

COLUMN_WIDTHS = {"A": 20, "B": 45, "C": 8, "D": 6, "E": 40, "F": 18, "G": 35, "H": 30}

HEADER_LABELS = [
    "Field ID", "Питання", "Strictness", "Фаза",
    "Значення", "Статус", "Коментар", "Джерело",
]

LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)


# ---------------------------------------------------------------------------
# YAML helpers
# ---------------------------------------------------------------------------

def _load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    if not isinstance(data, dict):
        raise TypeError(f"Expected mapping in {path}, got {type(data).__name__}")
    return data


# ---------------------------------------------------------------------------
# Index builders
# ---------------------------------------------------------------------------

def build_field_index(fields_data: dict) -> dict[str, dict]:
    return {f["field_id"]: f for f in fields_data["fields"]}


def build_section_maps(core_data: dict) -> tuple[dict[str, str], list[str]]:
    field_to_section: dict[str, str] = {}
    section_order: list[str] = []
    for section in core_data["sections"]:
        sid = section["id"]
        section_order.append(sid)
        for fid in section.get("fields", []):
            field_to_section[fid] = sid
    return field_to_section, section_order


def build_value_dicts(values_data: dict) -> dict[str, list[dict]]:
    return values_data.get("dictionaries", {})


def build_person_roles(role_data: dict) -> dict[str, dict]:
    persons: dict[str, dict] = {}
    for a in role_data["assignments"]:
        pid = a["person_id"]
        if pid in persons:
            persons[pid]["roles"].update(a["roles"])
        else:
            persons[pid] = {
                "label_uk": a.get("label_uk", pid),
                "roles": set(a["roles"]),
            }
    return persons


def assign_fields_to_persons(
    field_index: dict[str, dict],
    person_roles: dict[str, dict],
) -> tuple[dict[str, list[str]], list[str]]:
    role_to_person: dict[str, str] = {}
    for pid, info in person_roles.items():
        for role in info["roles"]:
            if role in role_to_person:
                raise ValueError(
                    f"Role '{role}' assigned to both "
                    f"'{role_to_person[role]}' and '{pid}'."
                )
            role_to_person[role] = pid

    person_fields: dict[str, list[str]] = {pid: [] for pid in person_roles}
    unassigned: list[str] = []

    for fid, field in field_index.items():
        owner_role = field["owner_role"]
        pid = role_to_person.get(owner_role)
        if pid:
            person_fields[pid].append(fid)
        else:
            unassigned.append(fid)
            print(
                f"WARNING: field '{fid}' (owner_role: {owner_role}) "
                "has no assigned person",
                file=sys.stderr,
            )

    return person_fields, unassigned


# ---------------------------------------------------------------------------
# Dropdown / comment helpers
# ---------------------------------------------------------------------------

def build_dropdown_items(dict_name: str, value_dicts: dict) -> list[str]:
    entries = value_dicts.get(dict_name, [])
    items = [f"{e['value_code']} \u2014 {e['label_uk']}" for e in entries]
    codes = {e["value_code"] for e in entries}
    if "tbd" not in codes:
        items.append("tbd \u2014 \u041d\u0435 \u0432\u0438\u0437\u043d\u0430\u0447\u0435\u043d\u043e")
    return items


def build_comment_text(dict_name: str, value_dicts: dict, field: dict) -> str:
    entries = value_dicts.get(dict_name, [])
    lines: list[str] = []
    for entry in entries:
        lines.append(f"{entry['value_code']} \u2014 {entry['label_uk']}")
        meaning = entry.get("meaning", "")
        if meaning:
            lines.append(f"  {meaning}")
        sel = entry.get("selection_rule", "")
        if sel:
            lines.append(f"  \u041e\u0431\u0438\u0440\u0430\u0439\u0442\u0435 \u044f\u043a\u0449\u043e: {sel}")
        lines.append("")

    field_sel = field.get("selection_rule", "")
    if field_sel:
        lines.append(f"\u041f\u0440\u0430\u0432\u0438\u043b\u043e \u0432\u0438\u0431\u043e\u0440\u0443: {field_sel}")
    impact = field.get("design_impact", [])
    if impact:
        lines.append(f"\u0412\u043f\u043b\u0438\u0432\u0430\u0454 \u043d\u0430: {', '.join(impact)}")
    return "\n".join(lines).strip()


# ---------------------------------------------------------------------------
# Sort key
# ---------------------------------------------------------------------------

def _field_sort_key(fid: str, field_to_section: dict[str, str]) -> tuple:
    section = field_to_section.get(fid, "")
    phase = PHASE_MAP.get(section, 99)
    return (phase, section, fid)


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _write_values_sheet(
    wb: Workbook,
    sorted_fields: list[str],
    field_index: dict[str, dict],
    value_dicts: dict,
) -> dict[str, str]:
    """Write dropdown values to a hidden _values sheet.

    Returns dict_name -> cell range string for use in DataValidation.
    Using cell ranges instead of inline comma-joined strings avoids the bug
    where a descriptive label containing a comma is split into two choices.
    """
    ws = wb.create_sheet("_values")
    ws.sheet_state = "hidden"

    # Collect unique dictionaries in field order
    used: list[str] = []
    seen: set[str] = set()
    for fid in sorted_fields:
        field = field_index[fid]
        if field.get("type", "enum") == "enum":
            ref = field.get("allowed_values_ref")
            if ref and ref not in seen:
                used.append(ref)
                seen.add(ref)

    dict_ranges: dict[str, str] = {}
    for col_idx, dict_name in enumerate(used, 1):
        items = build_dropdown_items(dict_name, value_dicts)
        col_letter = get_column_letter(col_idx)
        for row_idx, item in enumerate(items, 1):
            ws.cell(row=row_idx, column=col_idx, value=item)
        dict_ranges[dict_name] = (
            f"_values!${col_letter}$1:${col_letter}${len(items)}"
        )

    return dict_ranges


def write_xlsx(
    path: Path,
    person_id: str,
    label_uk: str,
    roles: set[str],
    field_ids: list[str],
    field_index: dict[str, dict],
    field_to_section: dict[str, str],
    value_dicts: dict,
    object_id: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "intake"

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # --- Title rows 1-5 ---
    roles_str = ", ".join(sorted(roles))
    today = date.today().isoformat()

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Intake Sheet \u2014 {label_uk} ({object_id})"
    ws["A1"].font = Font(bold=True, size=14)

    ws.merge_cells("A2:H2")
    ws["A2"] = f"\u0420\u043e\u043b\u0456: {roles_str}  |  \u0414\u0430\u0442\u0430 \u0433\u0435\u043d\u0435\u0440\u0430\u0446\u0456\u0457: {today}"

    ws.merge_cells("A3:H3")
    ws["A3"] = "\u0417\u0430\u043f\u043e\u0432\u043d\u0456\u0442\u044c \u043a\u043e\u043b\u043e\u043d\u043a\u0443 E. \u042f\u043a\u0449\u043e \u043d\u0435\u0432\u0456\u0434\u043e\u043c\u043e \u2014 \u043e\u0431\u0435\u0440\u0456\u0442\u044c tbd \u0443 \u043a\u043e\u043b\u043e\u043d\u0446\u0456 F."

    ws.merge_cells("A4:H4")
    ws["A4"] = f"\u0414\u0435\u0442\u0430\u043b\u044c\u043d\u0438\u0439 \u043e\u043f\u0438\u0441 \u043f\u043e\u043b\u0456\u0432: intake/generated/{person_id}.guide.md"

    # Row 5: empty separator

    # --- Header row 6 (frozen) ---
    HEADER_ROW = 6
    for col_idx, label in enumerate(HEADER_LABELS, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.protection = LOCKED
        cell.number_format = numbers.FORMAT_TEXT

    ws.freeze_panes = "A7"

    # --- Sort fields ---
    sorted_fields = sorted(
        field_ids,
        key=lambda fid: _field_sort_key(fid, field_to_section),
    )

    # --- Hidden _values sheet for dropdown ranges (avoids comma-in-label bug) ---
    dict_ranges = _write_values_sheet(wb, sorted_fields, field_index, value_dicts)

    # --- Status validation (shared across all rows) ---
    status_dv = DataValidation(
        type="list",
        formula1='"tbd,not_applicable"',
        allow_blank=True,
    )
    status_dv.prompt = "\u0417\u0430\u043b\u0438\u0448\u0442\u0435 \u043f\u043e\u0440\u043e\u0436\u043d\u0456\u043c \u044f\u043a\u0449\u043e \u0432\u0456\u0434\u043f\u043e\u0432\u0456\u0434\u044c \u0443 \u043a\u043e\u043b\u043e\u043d\u0446\u0456 E"
    status_dv.promptTitle = "\u0421\u0442\u0430\u0442\u0443\u0441"
    ws.add_data_validation(status_dv)

    # --- Data rows (7+) ---
    for row_offset, fid in enumerate(sorted_fields):
        row = HEADER_ROW + 1 + row_offset
        field = field_index[fid]
        section = field_to_section.get(fid, "")
        phase = PHASE_MAP.get(section, 0)
        strictness = field.get("strictness", "S1")
        fill = STRICTNESS_FILLS.get(strictness, STRICTNESS_FILLS["S1"])

        # A: Field ID (locked)
        c_a = ws.cell(row=row, column=1, value=fid)
        c_a.protection = LOCKED
        c_a.number_format = numbers.FORMAT_TEXT
        c_a.fill = fill

        # B: label_uk (locked)
        c_b = ws.cell(row=row, column=2, value=field.get("label_uk", fid))
        c_b.protection = LOCKED
        c_b.number_format = numbers.FORMAT_TEXT
        c_b.fill = fill

        # C: Strictness (locked)
        c_c = ws.cell(row=row, column=3, value=strictness)
        c_c.protection = LOCKED
        c_c.number_format = numbers.FORMAT_TEXT
        c_c.fill = fill

        # D: Phase (locked)
        c_d = ws.cell(row=row, column=4, value=phase)
        c_d.protection = LOCKED
        c_d.number_format = numbers.FORMAT_TEXT
        c_d.fill = fill

        # E: Value (editable — dropdown/text/integer depending on field type)
        c_e = ws.cell(row=row, column=5)
        c_e.protection = UNLOCKED
        c_e.number_format = numbers.FORMAT_TEXT
        c_e.fill = fill

        field_type = field.get("type", "enum")

        if field_type == "enum":
            dict_ref = field.get("allowed_values_ref")
            if dict_ref and dict_ref in dict_ranges:
                dv = DataValidation(
                    type="list",
                    formula1=dict_ranges[dict_ref],
                    allow_blank=True,
                )
                dv.prompt = field.get("label_uk", fid)
                dv.promptTitle = "\u041e\u0431\u0435\u0440\u0456\u0442\u044c \u0437\u043d\u0430\u0447\u0435\u043d\u043d\u044f"
                ws.add_data_validation(dv)
                dv.add(c_e)

                comment_text = build_comment_text(dict_ref, value_dicts, field)
                if comment_text:
                    c_e.comment = Comment(comment_text, "intake-generator")

        elif field_type == "string":
            dv = DataValidation(allow_blank=True)
            dv.prompt = field.get("purpose", field.get("label_uk", fid))
            dv.promptTitle = field.get("label_uk", fid)
            ws.add_data_validation(dv)
            dv.add(c_e)

        elif field_type == "integer":
            dv = DataValidation(
                type="whole",
                operator="greaterThanOrEqual",
                formula1="0",
                allow_blank=True,
            )
            dv.prompt = field.get("purpose", field.get("label_uk", fid))
            dv.promptTitle = field.get("label_uk", fid)
            ws.add_data_validation(dv)
            dv.add(c_e)

        # F: Status (editable — tbd / not_applicable)
        c_f = ws.cell(row=row, column=6)
        c_f.protection = UNLOCKED
        c_f.number_format = numbers.FORMAT_TEXT
        c_f.fill = fill
        status_dv.add(c_f)

        # G: Comment (editable)
        c_g = ws.cell(row=row, column=7)
        c_g.protection = UNLOCKED
        c_g.number_format = numbers.FORMAT_TEXT
        c_g.fill = fill

        # H: Source (editable)
        c_h = ws.cell(row=row, column=8)
        c_h.protection = UNLOCKED
        c_h.number_format = numbers.FORMAT_TEXT
        c_h.fill = fill

    # --- Sheet protection (lock A-D, unlock E-H) ---
    ws.protection.sheet = True
    ws.protection.password = ""

    # --- Reference sheet ---
    _write_reference_sheet(wb, sorted_fields, field_index, value_dicts)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_reference_sheet(
    wb: Workbook,
    sorted_fields: list[str],
    field_index: dict[str, dict],
    value_dicts: dict,
) -> None:
    ws = wb.create_sheet("_reference")
    ws.protection.sheet = True

    headers = ["Dictionary", "Code", "Label", "Meaning", "Selection Rule"]
    widths = [25, 25, 30, 60, 60]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
        ws.column_dimensions[chr(64 + col)].width = w

    used_dicts: list[str] = []
    seen: set[str] = set()
    for fid in sorted_fields:
        ref = field_index[fid].get("allowed_values_ref")
        if ref and ref not in seen:
            used_dicts.append(ref)
            seen.add(ref)

    row = 2
    for dict_name in used_dicts:
        for entry in value_dicts.get(dict_name, []):
            ws.cell(row=row, column=1, value=dict_name)
            ws.cell(row=row, column=2, value=entry["value_code"])
            ws.cell(row=row, column=3, value=entry.get("label_uk", ""))
            ws.cell(row=row, column=4, value=entry.get("meaning", ""))
            ws.cell(row=row, column=5, value=entry.get("selection_rule", ""))
            row += 1


# ---------------------------------------------------------------------------
# Guide.md writer
# ---------------------------------------------------------------------------

def write_guide_md(
    path: Path,
    person_id: str,
    label_uk: str,
    roles: set[str],
    field_ids: list[str],
    field_index: dict[str, dict],
    field_to_section: dict[str, str],
    value_dicts: dict,
    object_id: str,
) -> None:
    lines = [
        f"# Intake Guide \u2014 {label_uk} ({object_id})",
        "",
        f"**\u0420\u043e\u043b\u0456:** {', '.join(sorted(roles))}",
        f"**\u0414\u0430\u0442\u0430 \u0433\u0435\u043d\u0435\u0440\u0430\u0446\u0456\u0457:** {date.today().isoformat()}",
        "",
        "---",
        "",
    ]

    sorted_fields = sorted(
        field_ids,
        key=lambda fid: _field_sort_key(fid, field_to_section),
    )

    current_section = None
    for fid in sorted_fields:
        field = field_index[fid]
        section = field_to_section.get(fid, "unknown")
        phase = PHASE_MAP.get(section, 0)

        if section != current_section:
            current_section = section
            lines.append(f"## {section} (Phase {phase})")
            lines.append("")

        strictness = field.get("strictness", "S1")
        lines.append(f"### {fid}")
        lines.append(f"**{field.get('label_uk', fid)}** | {strictness}")
        lines.append("")

        purpose = field.get("purpose", "")
        if purpose:
            lines.append(f"**\u041f\u0440\u0438\u0437\u043d\u0430\u0447\u0435\u043d\u043d\u044f:** {purpose}")
            lines.append("")

        sel_rule = field.get("selection_rule", "")
        if sel_rule:
            lines.append(f"**\u041f\u0440\u0430\u0432\u0438\u043b\u043e \u0432\u0438\u0431\u043e\u0440\u0443:** {sel_rule}")
            lines.append("")

        interp = field.get("interpretation_rule", "")
        if interp:
            lines.append(f"**\u0406\u043d\u0442\u0435\u0440\u043f\u0440\u0435\u0442\u0430\u0446\u0456\u044f:** {interp}")
            lines.append("")

        impact = field.get("design_impact", [])
        if impact:
            lines.append(f"**\u0412\u043f\u043b\u0438\u0432\u0430\u0454 \u043d\u0430:** {', '.join(impact)}")
            lines.append("")

        dict_ref = field.get("allowed_values_ref")
        if dict_ref:
            entries = value_dicts.get(dict_ref, [])
            if entries:
                lines.append("**\u0417\u043d\u0430\u0447\u0435\u043d\u043d\u044f:**")
                lines.append("")
                for entry in entries:
                    lines.append(
                        f"- `{entry['value_code']}` \u2014 {entry['label_uk']}"
                    )
                    meaning = entry.get("meaning", "")
                    if meaning:
                        lines.append(f"  {meaning}")
                    entry_sel = entry.get("selection_rule", "")
                    if entry_sel:
                        lines.append(f"  *\u041e\u0431\u0438\u0440\u0430\u0439\u0442\u0435 \u044f\u043a\u0449\u043e: {entry_sel}*")
                lines.append("")

        lines.append("---")
        lines.append("")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate(workspace_path: Path, project_root: Path | None = None) -> dict[str, Any]:
    """Generate intake sheets for all persons in workspace.

    Returns summary dict for testing/reporting.
    """
    if project_root is None:
        project_root = Path(__file__).resolve().parents[2]

    fields_data = _load_yaml(
        project_root / "specs" / "dictionary" / "questionnaire_v2_fields.yaml"
    )
    values_data = _load_yaml(
        project_root / "specs" / "dictionary" / "questionnaire_v2_values.yaml"
    )
    core_data = _load_yaml(
        project_root / "specs" / "questionnaire" / "core_questionnaire_v2.yaml"
    )
    role_data = _load_yaml(workspace_path / "role_assignments.yaml")

    field_index = build_field_index(fields_data)
    field_to_section, _ = build_section_maps(core_data)
    value_dicts = build_value_dicts(values_data)
    person_roles = build_person_roles(role_data)
    object_id = role_data.get("object_id", workspace_path.name)

    person_fields, unassigned = assign_fields_to_persons(field_index, person_roles)

    summary: dict[str, Any] = {
        "object_id": object_id,
        "persons": {},
        "unassigned_fields": unassigned,
    }

    for pid, fids in person_fields.items():
        if not fids:
            continue

        info = person_roles[pid]
        xlsx_path = workspace_path / "intake" / "responses" / f"{pid}.xlsx"
        guide_path = workspace_path / "intake" / "generated" / f"{pid}.guide.md"

        write_xlsx(
            xlsx_path, pid, info["label_uk"], info["roles"],
            fids, field_index, field_to_section, value_dicts, object_id,
        )
        write_guide_md(
            guide_path, pid, info["label_uk"], info["roles"],
            fids, field_index, field_to_section, value_dicts, object_id,
        )

        summary["persons"][pid] = {
            "roles": sorted(info["roles"]),
            "field_count": len(fids),
            "fields": sorted(fids),
            "xlsx": str(xlsx_path),
            "guide": str(guide_path),
        }

    if unassigned:
        unassigned_path = workspace_path / "intake" / "responses" / "_unassigned.xlsx"
        unassigned_guide = workspace_path / "intake" / "generated" / "_unassigned.guide.md"
        write_xlsx(
            unassigned_path, "_unassigned", "Нерозподілені поля", set(),
            unassigned, field_index, field_to_section, value_dicts, object_id,
        )
        write_guide_md(
            unassigned_guide, "_unassigned", "Нерозподілені поля", set(),
            unassigned, field_index, field_to_section, value_dicts, object_id,
        )
        summary["unassigned_xlsx"] = str(unassigned_path)
        summary["unassigned_guide"] = str(unassigned_guide)

    return summary


def main() -> None:
    if len(sys.argv) < 2:
        print(
            "Usage: generate_intake_sheets.py <workspace_path>",
            file=sys.stderr,
        )
        sys.exit(1)

    workspace = Path(sys.argv[1]).resolve()
    if not workspace.exists():
        print(f"Workspace not found: {workspace}", file=sys.stderr)
        sys.exit(1)

    summary = generate(workspace)

    print(f"Generated intake sheets for {summary['object_id']}:")
    for pid, info in summary["persons"].items():
        print(f"  {pid}: {info['field_count']} fields \u2192 {info['xlsx']}")
    if summary["unassigned_fields"]:
        n = len(summary["unassigned_fields"])
        print(f"  _unassigned: {n} fields \u2192 {summary.get('unassigned_xlsx', 'N/A')}")


if __name__ == "__main__":
    main()
