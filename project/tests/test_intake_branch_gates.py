from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import yaml

from compiler.build_requirements_model import (
    build_requirements_model,
    validate_requirements_model,
)
from compiler.compile_graphs import compile_all_graphs, summarize_graph_bundle
from intake.compile_intake import compile_intake
from intake.generate_intake_sheets import generate
from model_utils import load_yaml
from run_pipeline import run_validators, summarize_validation

from conftest import GOLDEN_DATE

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCHEMA = (
    PROJECT_ROOT
    / "specs"
    / "requirements"
    / "object_requirements_v2.schema.yaml"
)
HAPPY_PATH = PROJECT_ROOT / "examples" / "sample_object_01"
STRESS_PATH = PROJECT_ROOT / "examples" / "sample_object_02"


def _build_label_map(values_sheet) -> dict[str, str]:
    labels: dict[str, str] = {}
    for col in range(1, values_sheet.max_column + 1):
        for row in range(1, values_sheet.max_row + 1):
            value = values_sheet.cell(row, col).value
            if value and " — " in str(value):
                labels[str(value).split(" — ")[0].strip()] = str(value)
    return labels


def _hydrate_generated_workbooks_from_responses(
    workspace: Path,
    exemplar_workspace: Path,
) -> None:
    for response_path in sorted(
        (exemplar_workspace / "intake" / "responses").glob("*.response.yaml")
    ):
        xlsx_path = workspace / "intake" / "responses" / (
            response_path.name.replace(".response.yaml", ".xlsx")
        )
        response = yaml.safe_load(response_path.read_text(encoding="utf-8")) or {}
        fields = response.get("fields", {})

        wb = load_workbook(xlsx_path)
        intake = wb["intake"]
        labels = _build_label_map(wb["_values"])

        for row in range(7, intake.max_row + 1):
            field_id = intake.cell(row, 1).value
            if field_id is None:
                continue
            field_id = str(field_id).strip()
            entry = fields.get(field_id, {})

            intake.cell(row, 5).value = None
            intake.cell(row, 6).value = None
            intake.cell(row, 7).value = None
            intake.cell(row, 8).value = None

            status = entry.get("status", "unanswered")
            value = entry.get("value")
            if status == "answered" and value is not None:
                intake.cell(row, 5).value = labels.get(str(value), str(value))
            elif status in {"tbd", "not_applicable"}:
                intake.cell(row, 6).value = status

            if entry.get("comment"):
                intake.cell(row, 7).value = entry["comment"]
            if entry.get("source_ref"):
                intake.cell(row, 8).value = entry["source_ref"]

        wb.save(xlsx_path)


def _count_field_statuses(all_fields: dict[str, dict[str, Any]]) -> dict[str, int]:
    counts = {"answered": 0, "tbd": 0, "unanswered": 0, "not_applicable": 0}
    for entry in all_fields.values():
        counts[entry["status"]] += 1
    return counts


def _run_pipeline(questionnaire_path: Path) -> dict[str, Any]:
    questionnaire = load_yaml(questionnaire_path)
    requirements, assumptions = build_requirements_model(questionnaire)
    validate_requirements_model(requirements, schema=load_yaml(SCHEMA))

    role_assignments = None
    auto_role_assignments = questionnaire_path.parent / "role_assignments.yaml"
    if auto_role_assignments.exists():
        role_assignments = load_yaml(auto_role_assignments)

    bundle = compile_all_graphs(requirements)
    graph_summary = summarize_graph_bundle(bundle)
    issues = run_validators(
        requirements,
        graph_summary,
        bundle,
        assumptions,
        role_assignments=role_assignments,
    )
    return {
        "requirements": requirements,
        "assumptions": assumptions,
        "issues": issues,
        "validation": summarize_validation(issues),
    }


def _materialize_roundtrip_workspace(tmp_path: Path, exemplar_workspace: Path) -> tuple[Path, dict[str, Any], dict[str, Any]]:
    workspace = tmp_path / exemplar_workspace.name
    workspace.mkdir()
    shutil.copy(
        exemplar_workspace / "role_assignments.yaml",
        workspace / "role_assignments.yaml",
    )

    generate(
        workspace,
        project_root=PROJECT_ROOT,
        generated_on=GOLDEN_DATE,
    )
    _hydrate_generated_workbooks_from_responses(workspace, exemplar_workspace)

    compile_result = compile_intake(
        workspace,
        project_root=PROJECT_ROOT,
        compiled_on=GOLDEN_DATE,
    )
    pipeline_result = _run_pipeline(workspace / "questionnaire.yaml")
    return workspace, compile_result, pipeline_result


def test_gate_a_happy_path_roundtrip(tmp_path):
    workspace, compile_result, pipeline_result = _materialize_roundtrip_workspace(
        tmp_path, HAPPY_PATH
    )

    assert compile_result["warnings"] == []
    assert _count_field_statuses(compile_result["all_fields"]) == {
        "answered": 41,
        "tbd": 0,
        "unanswered": 0,
        "not_applicable": 0,
    }
    assert (workspace / "reports" / "intake_status.yaml").exists()
    assert pipeline_result["validation"]["status"] == "ok"
    assert pipeline_result["validation"]["error_count"] == 0
    assert pipeline_result["assumptions"] == []
    assert (
        pipeline_result["requirements"]["metadata"]["resolved_archetype"]
        == "video_heavy_site"
    )


def test_gate_c_stress_path_partial_roundtrip(tmp_path):
    workspace, compile_result, pipeline_result = _materialize_roundtrip_workspace(
        tmp_path, STRESS_PATH
    )

    assert compile_result["warnings"] == []
    assert _count_field_statuses(compile_result["all_fields"]) == {
        "answered": 29,
        "tbd": 7,
        "unanswered": 5,
        "not_applicable": 0,
    }
    assert (workspace / "reports" / "intake_status.yaml").exists()
    assert pipeline_result["validation"]["status"] == "failed"
    assert pipeline_result["validation"]["error_count"] == 2
    assert (
        pipeline_result["requirements"]["metadata"]["resolved_archetype"]
        == "resilient_telemetry_site"
    )


def test_gate_d_stress_expected_failure_contract(tmp_path):
    _, _, pipeline_result = _materialize_roundtrip_workspace(tmp_path, STRESS_PATH)

    errors = [issue for issue in pipeline_result["issues"] if issue["severity"] == "error"]
    error_validators = {issue["validator"] for issue in errors}

    assert error_validators == {"resilience", "time"}
    assert len(errors) == 2
    assert any("redundancy_target" in issue["message"] for issue in errors)
    assert any("PTP" in issue["message"] for issue in errors)
    assert not any(
        issue["validator"] == "role_assignments"
        for issue in pipeline_result["issues"]
    )
