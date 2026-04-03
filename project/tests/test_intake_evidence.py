from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from intake.evidence_status import (
    EVIDENCE_SCHEMA_VERSION,
    _load_evidence_policy,
    build_evidence_status_from_snapshot,
    evidence_workspace,
    main as evidence_status_main,
)
from intake.review_packets import review_workspace
from intake.workspace_snapshot import build_workspace_snapshot
from model_utils import load_yaml

from conftest import GOLDEN_DATE, copy_workspace, find_evidence_field, find_review_item

PROJECT_ROOT = Path(__file__).resolve().parents[1]
HAPPY_PATH = PROJECT_ROOT / "examples" / "sample_object_01"
STRESS_PATH = PROJECT_ROOT / "examples" / "sample_object_02"


def _set_source_ref(workspace: Path, field_id: str, source_ref: str) -> None:
    responses_dir = workspace / "intake" / "responses"
    for workbook_path in sorted(responses_dir.glob("*.xlsx")):
        workbook = load_workbook(workbook_path)
        worksheet = workbook["intake"]
        for row in range(7, worksheet.max_row + 1):
            if worksheet.cell(row, 1).value == field_id:
                worksheet.cell(row, 8).value = source_ref
                workbook.save(workbook_path)
                return
    raise AssertionError(f"Field {field_id!r} not found in any intake workbook")


def _set_field_raw_value(workspace: Path, field_id: str, raw_value: str | None) -> None:
    responses_dir = workspace / "intake" / "responses"
    for workbook_path in sorted(responses_dir.glob("*.xlsx")):
        workbook = load_workbook(workbook_path)
        worksheet = workbook["intake"]
        for row in range(7, worksheet.max_row + 1):
            if worksheet.cell(row, 1).value == field_id:
                worksheet.cell(row, 5).value = raw_value
                worksheet.cell(row, 6).value = None
                workbook.save(workbook_path)
                return
    raise AssertionError(f"Field {field_id!r} not found in any intake workbook")


def test_evidence_status_sample01_is_advisory_only_and_writes_reports(tmp_path):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)

    result = evidence_workspace(
        workspace,
        project_root=PROJECT_ROOT,
        evidence_on=GOLDEN_DATE,
    )

    yaml_report = workspace / "reports" / "evidence_status.yaml"
    markdown_report = workspace / "reports" / "evidence_status.md"
    payload = load_yaml(yaml_report)
    control_required = find_evidence_field(result, "control_required")
    fat_required = find_evidence_field(result, "fat_required")
    sat_required = find_evidence_field(result, "sat_required")

    assert yaml_report.exists()
    assert markdown_report.exists()
    assert result["schema_version"] == EVIDENCE_SCHEMA_VERSION
    assert payload["schema_version"] == EVIDENCE_SCHEMA_VERSION
    assert result["metadata"]["project_stage"] == "concept"
    assert result["gate"]["mode"] == "advisory_only"
    assert result["gate"]["status"] == "passed"
    assert result["summary"]["selected_field_count"] > 0
    assert result["summary"]["advisory_gap_count"] > 0
    assert result["summary"]["blocking_eligible_count"] == 0
    assert result["summary"]["blocking_gap_count"] == 0
    assert all(field["blocking_eligible"] is False for field in result["fields"])
    assert all(field["blocking_gap"] is False for field in result["fields"])
    assert control_required["blocking_stage_allowed"] is False
    assert fat_required["blocking_eligible"] is False
    assert sat_required["blocking_eligible"] is False
    markdown_text = markdown_report.read_text(encoding="utf-8")
    assert "Blocking-eligible fields: 0" in markdown_report.read_text(encoding="utf-8")
    assert "Evidence gate mode: `advisory_only`" in markdown_text


def test_evidence_status_blocking_applies_only_to_allowlisted_fields(tmp_path):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)
    _set_field_raw_value(workspace, "project_stage", "detailed_design")

    result = evidence_workspace(
        workspace,
        project_root=PROJECT_ROOT,
        evidence_on=GOLDEN_DATE,
    )

    control_required = find_evidence_field(result, "control_required")
    fat_required = find_evidence_field(result, "fat_required")
    sat_required = find_evidence_field(result, "sat_required")

    assert result["gate"]["mode"] == "blocking"
    assert control_required["blocking_stage_allowed"] is True
    assert control_required["blocking_allowlisted"] is False
    assert control_required["blocking_eligible"] is False
    assert control_required["blocking_gap"] is False
    assert fat_required["blocking_allowlisted"] is True
    assert fat_required["blocking_eligible"] is True
    assert sat_required["blocking_allowlisted"] is True
    assert sat_required["blocking_eligible"] is True
    assert set(result["gate"]["blocking_eligible_field_ids"]) == {"fat_required", "sat_required"}


def test_evidence_policy_rejects_blocking_default_flag(tmp_path):
    policy_dir = tmp_path / "specs" / "evidence"
    policy_dir.mkdir(parents=True, exist_ok=True)
    (policy_dir / "evidence_policy.yaml").write_text(
        "\n".join(
            [
                "schema_version: 0.1.0",
                "defaults:",
                "  advisory_minimum_strength: structured_ref",
                "  blocking_minimum_strength: workspace_artifact",
                "  blocking_enforced: true",
                "field_overrides: {}",
                "",
            ]
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="blocking_enforced"):
        _load_evidence_policy(tmp_path)


def test_evidence_policy_rejects_blocking_fields_with_weaker_advisory_threshold(tmp_path):
    policy_dir = tmp_path / "specs" / "evidence"
    policy_dir.mkdir(parents=True, exist_ok=True)
    (policy_dir / "evidence_policy.yaml").write_text(
        "\n".join(
            [
                "schema_version: 0.1.0",
                "defaults:",
                "  advisory_minimum_strength: structured_ref",
                "  blocking_minimum_strength: workspace_artifact",
                "  blocking_allowed_stages:",
                "    - detailed_design",
                "field_overrides:",
                "  fat_required:",
                "    blocking_enforced: true",
                "",
            ]
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="fat_required"):
        _load_evidence_policy(tmp_path)


def test_evidence_status_derives_reference_structured_and_workspace_artifact(tmp_path):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)
    _set_source_ref(workspace, "control_required", "site survey notebook 2026-04-02")
    _set_source_ref(workspace, "oob_required", "path=evidence/oob_basis.md")
    _set_source_ref(workspace, "sat_required", "path=evidence/sat_basis.md")
    artifact_dir = workspace / "evidence"
    artifact_dir.mkdir(parents=True, exist_ok=True)
    (artifact_dir / "sat_basis.md").write_text("SAT basis\n", encoding="utf-8")

    result = evidence_workspace(
        workspace,
        project_root=PROJECT_ROOT,
        evidence_on=GOLDEN_DATE,
    )

    control_required = find_evidence_field(result, "control_required")
    oob_required = find_evidence_field(result, "oob_required")
    sat_required = find_evidence_field(result, "sat_required")

    assert control_required["evidence_strength"] == "reference_only"
    assert control_required["advisory_gap"] is True
    assert control_required["gap_reason"] == "weak evidence: reference_only < structured_ref"
    assert control_required["review_routing_required"] is False

    assert oob_required["evidence_strength"] == "structured_ref"
    assert oob_required["advisory_gap"] is False
    assert oob_required["workspace_artifacts"] == []

    assert sat_required["evidence_strength"] == "workspace_artifact"
    assert sat_required["advisory_gap"] is False
    assert sat_required["workspace_artifacts"] == ["evidence/sat_basis.md"]


@pytest.mark.parametrize("project_stage", ["basic_design", "unexpected_stage", None])
def test_evidence_status_stays_advisory_outside_blocking_stage_matrix(tmp_path, project_stage):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)
    snapshot = build_workspace_snapshot(
        workspace,
        project_root=PROJECT_ROOT,
        snapshot_on=GOLDEN_DATE,
        write_pipeline_outputs=False,
    )
    snapshot["metadata"]["project_stage"] = project_stage

    result = build_evidence_status_from_snapshot(
        snapshot,
        project_root=PROJECT_ROOT,
    )

    fat_required = find_evidence_field(result, "fat_required")
    sat_required = find_evidence_field(result, "sat_required")

    assert result["gate"]["mode"] == "advisory_only"
    assert result["gate"]["status"] == "passed"
    assert result["summary"]["blocking_eligible_count"] == 0
    assert result["summary"]["blocking_gap_count"] == 0
    assert result["gate"]["blocking_eligible_field_ids"] == []
    assert fat_required["blocking_stage_allowed"] is False
    assert sat_required["blocking_stage_allowed"] is False


def test_review_packets_include_evidence_gap_for_unresolved_selected_field(tmp_path):
    workspace = copy_workspace(tmp_path, STRESS_PATH)

    result = review_workspace(
        workspace,
        project_root=PROJECT_ROOT,
        review_on=GOLDEN_DATE,
    )

    item = find_review_item(result, source_kind="evidence_gap", source_key="control_required")
    assert item["routing_state"] == "assigned"
    assert item["primary_role"] == "process_engineer"
    assert item["primary_person"] == "sample2_process_telemetry"
    assert item["review_item_id"] == "sample_object_02.evidence_gap.control_required.process_engineer"
    assert item["blocking_eligible"] is False
    assert item["blocking_gap"] is False
    assert "missing evidence" in item["review_reasons"]


@pytest.mark.parametrize("project_stage", ["detailed_design", "build_commission"])
def test_evidence_status_applies_blocking_at_allowed_stages(tmp_path, project_stage):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)
    _set_field_raw_value(workspace, "project_stage", project_stage)

    result = evidence_workspace(
        workspace,
        project_root=PROJECT_ROOT,
        evidence_on=GOLDEN_DATE,
    )

    control_required = find_evidence_field(result, "control_required")
    sat_required = find_evidence_field(result, "sat_required")
    fat_required = find_evidence_field(result, "fat_required")

    assert result["metadata"]["project_stage"] == project_stage
    assert result["gate"]["mode"] == "blocking"
    assert result["gate"]["status"] == "failed"
    assert result["summary"]["blocking_eligible_count"] == 2
    assert result["summary"]["blocking_gap_count"] == 2
    assert control_required["blocking_eligible"] is False
    assert fat_required["blocking_eligible"] is True
    assert fat_required["blocking_gap"] is True
    assert sat_required["blocking_eligible"] is True
    assert sat_required["blocking_gap"] is True


def test_evidence_status_enforces_workspace_artifact_as_minimum_blocking_strength(tmp_path):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)
    _set_field_raw_value(workspace, "project_stage", "detailed_design")
    _set_source_ref(workspace, "fat_required", "path=evidence/fat_basis.md")
    _set_source_ref(workspace, "sat_required", "path=evidence/sat_basis.md")
    artifact_dir = workspace / "evidence"
    artifact_dir.mkdir(parents=True, exist_ok=True)
    (artifact_dir / "sat_basis.md").write_text("SAT basis\n", encoding="utf-8")

    result = evidence_workspace(
        workspace,
        project_root=PROJECT_ROOT,
        evidence_on=GOLDEN_DATE,
    )

    fat_required = find_evidence_field(result, "fat_required")
    sat_required = find_evidence_field(result, "sat_required")

    assert fat_required["evidence_strength"] == "structured_ref"
    assert fat_required["blocking_eligible"] is True
    assert fat_required["blocking_gap"] is True
    assert fat_required["blocking_reason"] == (
        "blocking evidence gap: structured_ref < workspace_artifact"
    )
    assert fat_required["review_routing_required"] is True

    assert sat_required["evidence_strength"] == "workspace_artifact"
    assert sat_required["blocking_eligible"] is True
    assert sat_required["blocking_gap"] is False
    assert sat_required["blocking_reason"] is None

    assert result["summary"]["blocking_gap_count"] == 1
    assert result["gate"]["blocking_gap_field_ids"] == ["fat_required"]


def test_evidence_status_gate_passes_when_all_blocking_evidence_is_present(tmp_path):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)
    _set_field_raw_value(workspace, "project_stage", "detailed_design")
    _set_source_ref(workspace, "fat_required", "path=evidence/fat_basis.md")
    _set_source_ref(workspace, "sat_required", "path=evidence/sat_basis.md")
    artifact_dir = workspace / "evidence"
    artifact_dir.mkdir(parents=True, exist_ok=True)
    (artifact_dir / "fat_basis.md").write_text("FAT basis\n", encoding="utf-8")
    (artifact_dir / "sat_basis.md").write_text("SAT basis\n", encoding="utf-8")

    result = evidence_workspace(
        workspace,
        project_root=PROJECT_ROOT,
        evidence_on=GOLDEN_DATE,
    )

    fat_required = find_evidence_field(result, "fat_required")
    sat_required = find_evidence_field(result, "sat_required")

    assert result["gate"]["mode"] == "blocking"
    assert result["gate"]["status"] == "passed"
    assert result["summary"]["blocking_eligible_count"] == 2
    assert result["summary"]["blocking_gap_count"] == 0
    assert result["gate"]["blocking_gap_field_ids"] == []
    assert fat_required["blocking_gap"] is False
    assert sat_required["blocking_gap"] is False


def test_review_packets_mark_blocking_evidence_gaps_in_reviewer_packet(tmp_path):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)
    _set_field_raw_value(workspace, "project_stage", "detailed_design")
    _set_source_ref(workspace, "sat_required", "path=evidence/sat_basis.md")
    artifact_dir = workspace / "evidence"
    artifact_dir.mkdir(parents=True, exist_ok=True)
    (artifact_dir / "sat_basis.md").write_text("SAT basis\n", encoding="utf-8")

    result = review_workspace(
        workspace,
        project_root=PROJECT_ROOT,
        review_on=GOLDEN_DATE,
    )

    item = find_review_item(result, source_kind="evidence_gap", source_key="fat_required")
    assert item["blocking_eligible"] is True
    assert item["blocking_gap"] is True
    assert item["priority"] == "critical"
    assert "blocking evidence gap" in item["review_reasons"]
    assert item["primary_person"] is not None

    packet_text = (
        workspace / "reports" / f"review_packet.{item['primary_person']}.md"
    ).read_text(encoding="utf-8")
    assert "Blocking gap: yes" in packet_text
    assert "Blocking reason: blocking evidence gap: missing evidence" in packet_text


def test_evidence_cli_exits_non_zero_after_writing_blocking_reports(
    tmp_path,
    monkeypatch,
    capsys,
):
    workspace = copy_workspace(tmp_path, HAPPY_PATH)
    _set_field_raw_value(workspace, "project_stage", "detailed_design")
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "evidence_status.py",
            str(workspace),
            "--date",
            GOLDEN_DATE.isoformat(),
        ],
    )

    with pytest.raises(SystemExit) as exc:
        evidence_status_main()

    assert exc.value.code == 1
    assert (workspace / "reports" / "evidence_status.yaml").exists()
    assert (workspace / "reports" / "evidence_status.md").exists()
    payload = load_yaml(workspace / "reports" / "evidence_status.yaml")
    assert payload["gate"]["status"] == "failed"
    assert "gate_status: failed" in capsys.readouterr().out
