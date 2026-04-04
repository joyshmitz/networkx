"""Tests for intake sheet generation (generate_intake_sheets.py).

Covers: fields per person, dropdown labels, cell comments, strictness colors,
field types (enum/string/integer), unassigned xlsx, guide.md, reference sheet.
"""
from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from openpyxl import load_workbook
import yaml

from network_methodology_sandbox.intake.generate_intake_sheets import (
    assign_fields_to_persons,
    build_field_index,
    build_person_roles,
    build_section_maps,
    build_value_dicts,
    generate,
    _load_yaml,
    PHASE_MAP,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_WORKSPACE = PROJECT_ROOT / "examples" / "sample_object_01"
SAMPLE_WORKSPACE_02 = PROJECT_ROOT / "examples" / "sample_object_02"

from conftest import GOLDEN_DATE as FIXED_DATE


@pytest.fixture(scope="module")
def summary(tmp_path_factory):
    """Run generate once into a temp workspace and return summary."""
    tmp = tmp_path_factory.mktemp("intake")
    workspace = tmp / "sample_object_01"
    workspace.mkdir()

    # Copy role_assignments.yaml
    import shutil
    shutil.copy(
        SAMPLE_WORKSPACE / "role_assignments.yaml",
        workspace / "role_assignments.yaml",
    )

    return generate(
        workspace,
        project_root=PROJECT_ROOT,
        generated_on=FIXED_DATE,
    )


@pytest.fixture(scope="module")
def workspace_path(summary):
    """Return workspace path from summary."""
    # Derive from first person's xlsx path
    first_person = next(iter(summary["persons"].values()))
    return Path(first_person["xlsx"]).parents[2]


@pytest.fixture(scope="module")
def summary_sample02(tmp_path_factory):
    """Run generate for sample_object_02 into a temp workspace."""
    tmp = tmp_path_factory.mktemp("intake02")
    workspace = tmp / "sample_object_02"
    workspace.mkdir()

    import shutil
    shutil.copy(
        SAMPLE_WORKSPACE_02 / "role_assignments.yaml",
        workspace / "role_assignments.yaml",
    )

    return generate(
        workspace,
        project_root=PROJECT_ROOT,
        generated_on=FIXED_DATE,
    )


# ---------------------------------------------------------------------------
# Field distribution
# ---------------------------------------------------------------------------


class TestFieldDistribution:
    def test_total_fields_equals_41(self, summary):
        total = sum(p["field_count"] for p in summary["persons"].values())
        total += len(summary["unassigned_fields"])
        assert total == 41

    def test_no_duplicate_fields(self, summary):
        all_fields = []
        for p in summary["persons"].values():
            all_fields.extend(p["fields"])
        all_fields.extend(summary["unassigned_fields"])
        assert len(all_fields) == len(set(all_fields))

    def test_sample_arch_has_5_fields(self, summary):
        assert summary["persons"]["sample_arch"]["field_count"] == 5

    def test_sample_arch_roles(self, summary):
        assert summary["persons"]["sample_arch"]["roles"] == [
            "network_engineer", "ot_architect",
        ]

    def test_unassigned_has_4_fields(self, summary):
        assert len(summary["unassigned_fields"]) == 4

    def test_unassigned_fields_are_process_iiot(self, summary):
        expected = {
            "telemetry_required", "control_required",
            "iiot_required", "local_archiving_required",
        }
        assert set(summary["unassigned_fields"]) == expected

    def test_five_persons_generated(self, summary):
        assert len(summary["persons"]) == 5


class TestFieldDistributionSample02:
    def test_all_41_fields_assigned(self, summary_sample02):
        total = sum(p["field_count"] for p in summary_sample02["persons"].values())
        assert total == 41

    def test_no_unassigned_fields(self, summary_sample02):
        assert summary_sample02["unassigned_fields"] == []

    def test_six_persons_generated(self, summary_sample02):
        assert len(summary_sample02["persons"]) == 6


# ---------------------------------------------------------------------------
# Excel structure
# ---------------------------------------------------------------------------


class TestExcelStructure:
    def test_sheets_present(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        assert wb.sheetnames == ["intake", "_values", "_reference"]

    def test_title_row(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        title = ws["A1"].value
        assert "sample_object_01" in title
        assert "Intake Sheet" in title

    def test_header_row_6(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        headers = [ws.cell(6, c).value for c in range(1, 9)]
        assert headers[0] == "Field ID"
        assert headers[4] == "Значення"

    def test_freeze_panes(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        assert ws.freeze_panes == "A7"

    def test_data_starts_row_7(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        assert ws.cell(7, 1).value is not None

    def test_explicit_date_written_to_xlsx_and_guide(self, tmp_path):
        import shutil

        workspace = tmp_path / "dated_workspace"
        workspace.mkdir()
        shutil.copy(
            SAMPLE_WORKSPACE / "role_assignments.yaml",
            workspace / "role_assignments.yaml",
        )

        dated_summary = generate(
            workspace,
            project_root=PROJECT_ROOT,
            generated_on=FIXED_DATE,
        )

        xlsx = dated_summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        assert FIXED_DATE.isoformat() in str(ws["A2"].value)

        guide_path = Path(dated_summary["persons"]["sample_arch"]["guide"])
        guide_text = guide_path.read_text(encoding="utf-8")
        assert f"**Дата генерації:** {FIXED_DATE.isoformat()}" in guide_text

    def test_fixed_date_generation_is_byte_stable(self, tmp_path):
        import shutil

        workspaces: list[Path] = []
        for name in ("first", "second"):
            workspace = tmp_path / name
            workspace.mkdir()
            shutil.copy(
                SAMPLE_WORKSPACE / "role_assignments.yaml",
                workspace / "role_assignments.yaml",
            )
            generate(
                workspace,
                project_root=PROJECT_ROOT,
                generated_on=FIXED_DATE,
            )
            workspaces.append(workspace)

        left, right = workspaces
        rel_paths = sorted(
            path.relative_to(left)
            for path in (left / "intake").rglob("*")
            if path.is_file()
        )
        for rel_path in rel_paths:
            left_hash = hashlib.sha256((left / rel_path).read_bytes()).hexdigest()
            right_hash = hashlib.sha256((right / rel_path).read_bytes()).hexdigest()
            assert left_hash == right_hash, rel_path.as_posix()


# ---------------------------------------------------------------------------
# Dropdowns and comments
# ---------------------------------------------------------------------------


class TestDropdowns:
    def test_enum_field_has_dropdown_via_values_sheet(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        for row in range(7, 20):
            if ws.cell(row, 1).value == "redundancy_target":
                found = False
                for dv in ws.data_validations.dataValidation:
                    if dv.type == "list" and f"E{row}" in str(dv.sqref):
                        found = True
                        assert "_values!" in dv.formula1
                        break
                assert found, "No dropdown found for redundancy_target"
                return
        pytest.fail("redundancy_target not found in sample_arch xlsx")

    def test_values_sheet_hidden(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        assert wb["_values"].sheet_state == "hidden"

    def test_comma_in_label_not_split(self, summary):
        """hybrid_controlled label contains a comma — must be one cell, not two."""
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws_val = wb["_values"]
        found = False
        for row in range(1, 20):
            for col in range(1, 20):
                val = ws_val.cell(row, col).value
                if val and "hybrid_controlled" in str(val):
                    assert "Частково спільний, частково розділений" in val
                    found = True
                    break
            if found:
                break
        assert found, "hybrid_controlled not found in _values sheet"

    def test_enum_field_has_cell_comment(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        for row in range(7, 20):
            if ws.cell(row, 1).value == "redundancy_target":
                comment = ws.cell(row, 5).comment
                assert comment is not None
                assert "none" in comment.text
                assert "Впливає на:" in comment.text
                return
        pytest.fail("redundancy_target not found")


# ---------------------------------------------------------------------------
# Strictness colors
# ---------------------------------------------------------------------------


class TestStrictnessColors:
    def test_s4_red_fill(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        for row in range(7, 20):
            if ws.cell(row, 3).value == "S4":
                fill = ws.cell(row, 1).fill
                # FFCCCC is the expected red fill
                assert "FFCCCC" in str(fill.start_color.rgb)
                return
        pytest.fail("No S4 field found")

    def test_s3_orange_fill(self, summary):
        xlsx = summary["persons"]["sample_ops_sec"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        for row in range(7, 30):
            if ws.cell(row, 3).value == "S3":
                fill = ws.cell(row, 1).fill
                assert "FFE0CC" in str(fill.start_color.rgb)
                return
        pytest.fail("No S3 field found")


# ---------------------------------------------------------------------------
# Field types
# ---------------------------------------------------------------------------


class TestFieldTypes:
    def test_string_field_no_list_validation(self, summary):
        """object_id is type=string, should not have list validation."""
        xlsx = summary["persons"]["sample_pm_owner"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        for row in range(7, 20):
            if ws.cell(row, 1).value == "object_id":
                for dv in ws.data_validations.dataValidation:
                    if f"E{row}" in str(dv.sqref):
                        assert dv.type != "list", "string field should not have list validation"
                return
        pytest.fail("object_id not found in sample_pm_owner xlsx")

    def test_integer_field_whole_validation(self, summary):
        """growth_horizon_months is type=integer, should have whole number validation."""
        xlsx = summary["persons"]["sample_pm_owner"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        for row in range(7, 20):
            if ws.cell(row, 1).value == "growth_horizon_months":
                found = False
                for dv in ws.data_validations.dataValidation:
                    if f"E{row}" in str(dv.sqref) and dv.type == "whole":
                        found = True
                        break
                assert found, "integer field should have whole number validation"
                return
        pytest.fail("growth_horizon_months not found")


# ---------------------------------------------------------------------------
# Protection
# ---------------------------------------------------------------------------


class TestProtection:
    def test_sheet_protected(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        assert ws.protection.sheet is True

    def test_column_a_locked(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        assert ws.cell(7, 1).protection.locked is True

    def test_column_e_unlocked(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        assert ws.cell(7, 5).protection.locked is False

    def test_column_g_unlocked(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        assert ws.cell(7, 7).protection.locked is False


# ---------------------------------------------------------------------------
# Text number format
# ---------------------------------------------------------------------------


class TestNumberFormat:
    def test_all_cells_text_format(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        for row in range(7, 12):
            for col in range(1, 9):
                fmt = ws.cell(row, col).number_format
                assert fmt == "@", (
                    f"Cell ({row},{col}) has format '{fmt}', expected '@'"
                )


# ---------------------------------------------------------------------------
# Unassigned xlsx
# ---------------------------------------------------------------------------


class TestUnassigned:
    def test_unassigned_xlsx_created(self, summary):
        assert "unassigned_xlsx" in summary
        assert Path(summary["unassigned_xlsx"]).exists()

    def test_unassigned_has_4_rows(self, summary):
        wb = load_workbook(summary["unassigned_xlsx"])
        ws = wb["intake"]
        count = 0
        for row in range(7, 20):
            if ws.cell(row, 1).value is None:
                break
            count += 1
        assert count == 4

    def test_unassigned_title_mentions_unassigned(self, summary):
        wb = load_workbook(summary["unassigned_xlsx"])
        ws = wb["intake"]
        assert "Нерозподілені" in ws["A1"].value

    def test_unassigned_guide_created(self, summary):
        assert "unassigned_guide" in summary
        assert Path(summary["unassigned_guide"]).exists()

    def test_unassigned_guide_has_fields(self, summary):
        text = Path(summary["unassigned_guide"]).read_text(encoding="utf-8")
        assert "telemetry_required" in text


# ---------------------------------------------------------------------------
# Guide.md
# ---------------------------------------------------------------------------


class TestGuide:
    def test_guide_files_created(self, summary):
        for pid, info in summary["persons"].items():
            assert Path(info["guide"]).exists(), f"Guide missing for {pid}"

    def test_guide_contains_field_section(self, summary):
        guide = Path(summary["persons"]["sample_arch"]["guide"])
        text = guide.read_text(encoding="utf-8")
        assert "## external_transport" in text
        assert "### wan_required" in text

    def test_guide_contains_values(self, summary):
        guide = Path(summary["persons"]["sample_arch"]["guide"])
        text = guide.read_text(encoding="utf-8")
        assert "`no_spof`" in text


# ---------------------------------------------------------------------------
# Reference sheet
# ---------------------------------------------------------------------------


class TestReferenceSheet:
    def test_reference_sheet_protected(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["_reference"]
        assert ws.protection.sheet is True

    def test_reference_has_entries(self, summary):
        xlsx = summary["persons"]["sample_arch"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["_reference"]
        assert ws.cell(2, 1).value is not None


# ---------------------------------------------------------------------------
# Phase ordering
# ---------------------------------------------------------------------------


class TestPhaseOrdering:
    def test_fields_sorted_by_phase(self, summary):
        """All fields in a person's xlsx should be sorted by phase."""
        xlsx = summary["persons"]["sample_ops_sec"]["xlsx"]
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        phases = []
        for row in range(7, 30):
            val = ws.cell(row, 4).value
            if val is None:
                break
            phases.append(int(val))
        assert phases == sorted(phases), f"Phases not sorted: {phases}"


# ---------------------------------------------------------------------------
# Edge case: duplicate role → error
# ---------------------------------------------------------------------------


class TestOwnershipValidation:
    def test_duplicate_role_raises(self):
        person_roles = {
            "person_a": {"label_uk": "A", "roles": {"ot_architect"}},
            "person_b": {"label_uk": "B", "roles": {"ot_architect"}},
        }
        fields_data = _load_yaml(
            PROJECT_ROOT / "specs" / "dictionary" / "questionnaire_v2_fields.yaml"
        )
        field_index = build_field_index(fields_data)
        with pytest.raises(ValueError, match="Role.*assigned to both"):
            assign_fields_to_persons(field_index, person_roles)


# ---------------------------------------------------------------------------
# Preserve responses
# ---------------------------------------------------------------------------


class TestPreserveResponses:
    @staticmethod
    def _find_row(ws, field_id: str) -> int:
        for row in range(7, ws.max_row + 1):
            if ws.cell(row, 1).value == field_id:
                return row
        raise AssertionError(f"{field_id} not found")

    def test_preserve_responses_keeps_existing_cells(self, tmp_path):
        import shutil

        workspace = tmp_path / "preserve_workspace"
        workspace.mkdir()
        shutil.copy(
            SAMPLE_WORKSPACE / "role_assignments.yaml",
            workspace / "role_assignments.yaml",
        )

        generate(workspace, project_root=PROJECT_ROOT, generated_on=FIXED_DATE)

        xlsx = workspace / "intake" / "responses" / "sample_pm_owner.xlsx"
        wb = load_workbook(xlsx)
        ws = wb["intake"]
        object_name_row = self._find_row(ws, "object_name")
        evidence_row = self._find_row(ws, "acceptance_evidence_class")
        ws.cell(object_name_row, 5).value = "Preserved Object Name"
        ws.cell(object_name_row, 7).value = "captured from workshop"
        ws.cell(object_name_row, 8).value = "meeting-notes-01"
        ws.cell(evidence_row, 6).value = "tbd"
        wb.save(xlsx)

        regenerate_summary = generate(
            workspace,
            project_root=PROJECT_ROOT,
            generated_on=FIXED_DATE,
            preserve_responses=True,
        )

        assert regenerate_summary["preserved_field_count"] >= 2

        regenerated = load_workbook(xlsx)
        ws = regenerated["intake"]
        object_name_row = self._find_row(ws, "object_name")
        evidence_row = self._find_row(ws, "acceptance_evidence_class")
        assert ws.cell(object_name_row, 5).value == "Preserved Object Name"
        assert ws.cell(object_name_row, 7).value == "captured from workshop"
        assert ws.cell(object_name_row, 8).value == "meeting-notes-01"
        assert ws.cell(evidence_row, 6).value == "tbd"

    def test_preserve_responses_moves_field_to_new_owner(self, tmp_path):
        import shutil

        workspace = tmp_path / "reassigned_workspace"
        workspace.mkdir()
        role_assignments_path = workspace / "role_assignments.yaml"
        shutil.copy(SAMPLE_WORKSPACE / "role_assignments.yaml", role_assignments_path)

        generate(workspace, project_root=PROJECT_ROOT, generated_on=FIXED_DATE)

        unassigned_xlsx = workspace / "intake" / "responses" / "_unassigned.xlsx"
        wb = load_workbook(unassigned_xlsx)
        ws = wb["intake"]
        telemetry_row = self._find_row(ws, "telemetry_required")
        ws.cell(telemetry_row, 6).value = "tbd"
        ws.cell(telemetry_row, 7).value = "waiting for process engineer"
        ws.cell(telemetry_row, 8).value = "ticket-42"
        wb.save(unassigned_xlsx)

        role_data = yaml.safe_load(role_assignments_path.read_text(encoding="utf-8"))
        for assignment in role_data["assignments"]:
            if assignment["person_id"] == "sample_field":
                assignment["roles"].append("process_engineer")
                break
        role_assignments_path.write_text(
            yaml.safe_dump(role_data, sort_keys=False, allow_unicode=True),
            encoding="utf-8",
        )

        regenerate_summary = generate(
            workspace,
            project_root=PROJECT_ROOT,
            generated_on=FIXED_DATE,
            preserve_responses=True,
        )

        assert regenerate_summary["unassigned_fields"] == ["iiot_required"]

        sample_field_xlsx = workspace / "intake" / "responses" / "sample_field.xlsx"
        regenerated = load_workbook(sample_field_xlsx)
        ws = regenerated["intake"]
        telemetry_row = self._find_row(ws, "telemetry_required")
        assert ws.cell(telemetry_row, 6).value == "tbd"
        assert ws.cell(telemetry_row, 7).value == "waiting for process engineer"
        assert ws.cell(telemetry_row, 8).value == "ticket-42"
