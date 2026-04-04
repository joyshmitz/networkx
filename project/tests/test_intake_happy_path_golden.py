from __future__ import annotations

import hashlib
from pathlib import Path
import shutil
from tempfile import NamedTemporaryFile
import xml.etree.ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import load_workbook
import yaml

from network_methodology_sandbox.intake.compile_intake import compile_intake
from network_methodology_sandbox.intake.generate_intake_sheets import generate


from conftest import GOLDEN_DATETIME_UTC

PROJECT_ROOT = Path(__file__).resolve().parents[1]
HAPPY_PATH = PROJECT_ROOT / "examples" / "sample_object_01"
FIXED_DATE = GOLDEN_DATETIME_UTC
FIXED_DATE_W3CDTF = FIXED_DATE.strftime("%Y-%m-%dT%H:%M:%SZ")
FIXED_ZIP_DATETIME = (FIXED_DATE.year, FIXED_DATE.month, FIXED_DATE.day, 0, 0, 0)

HAPPY_PATH_GOLDEN_FILES = (
    "questionnaire.yaml",
    "intake/generated/_unassigned.guide.md",
    "intake/generated/sample_arch.guide.md",
    "intake/generated/sample_field.guide.md",
    "intake/generated/sample_ops_sec.guide.md",
    "intake/generated/sample_pm_owner.guide.md",
    "intake/generated/sample_power_video.guide.md",
    "intake/responses/_unassigned.response.yaml",
    "intake/responses/_unassigned.xlsx",
    "intake/responses/sample_arch.response.yaml",
    "intake/responses/sample_arch.xlsx",
    "intake/responses/sample_field.response.yaml",
    "intake/responses/sample_field.xlsx",
    "intake/responses/sample_ops_sec.response.yaml",
    "intake/responses/sample_ops_sec.xlsx",
    "intake/responses/sample_pm_owner.response.yaml",
    "intake/responses/sample_pm_owner.xlsx",
    "intake/responses/sample_power_video.response.yaml",
    "intake/responses/sample_power_video.xlsx",
    "reports/intake_status.md",
    "reports/intake_status.yaml",
)


def _collect_tracked_happy_path_golden_files() -> tuple[str, ...]:
    return tuple(
        sorted(
            str(rel_path)
            for path in HAPPY_PATH.rglob("*")
            for rel_path in [path.relative_to(HAPPY_PATH)]
            if path.is_file()
            and (
                rel_path == Path("questionnaire.yaml")
                or rel_path.match("intake/generated/*.guide.md")
                or rel_path.match("intake/responses/*.xlsx")
                or rel_path.match("intake/responses/*.response.yaml")
                or rel_path == Path("reports/intake_status.yaml")
                or rel_path == Path("reports/intake_status.md")
            )
        )
    )


def _normalize_xlsx_core_properties(path: Path) -> None:
    namespaces = {
        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
        "dcterms": "http://purl.org/dc/terms/",
    }
    ET.register_namespace("cp", namespaces["cp"])
    ET.register_namespace("dc", "http://purl.org/dc/elements/1.1/")
    ET.register_namespace("dcterms", namespaces["dcterms"])
    ET.register_namespace("xsi", "http://www.w3.org/2001/XMLSchema-instance")

    with ZipFile(path, "r") as src:
        core_xml = src.read("docProps/core.xml")
        root = ET.fromstring(core_xml)
        created = root.find("dcterms:created", namespaces)
        modified = root.find("dcterms:modified", namespaces)
        assert created is not None
        assert modified is not None
        created.text = FIXED_DATE_W3CDTF
        modified.text = FIXED_DATE_W3CDTF

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            temp_path = Path(tmp_file.name)

        try:
            with ZipFile(temp_path, "w", compression=ZIP_DEFLATED) as dst:
                for info in src.infolist():
                    data = src.read(info.filename)
                    if info.filename == "docProps/core.xml":
                        data = ET.tostring(root, encoding="utf-8", xml_declaration=False)
                    stable_info = ZipInfo(info.filename, date_time=FIXED_ZIP_DATETIME)
                    stable_info.compress_type = ZIP_DEFLATED
                    stable_info.comment = info.comment
                    stable_info.extra = info.extra
                    stable_info.internal_attr = 0
                    stable_info.external_attr = 0
                    stable_info.create_system = 0
                    stable_info.flag_bits = info.flag_bits
                    dst.writestr(stable_info, data)
            temp_path.replace(path)
        finally:
            if temp_path.exists():
                temp_path.unlink()


def _fill_generated_workbooks_from_questionnaire(workspace: Path) -> None:
    questionnaire = yaml.safe_load(
        (HAPPY_PATH / "questionnaire.yaml").read_text(encoding="utf-8")
    )
    flat_answers: dict[str, object] = {}
    for section, values in questionnaire.items():
        if isinstance(values, dict) and section not in {"version", "known_unknowns"}:
            flat_answers.update(values)

    for xlsx in sorted((workspace / "intake" / "responses").glob("*.xlsx")):
        wb = load_workbook(xlsx)
        wb.properties.created = FIXED_DATE
        wb.properties.modified = FIXED_DATE
        intake = wb["intake"]
        values_sheet = wb["_values"]

        labels: dict[str, str] = {}
        for col in range(1, values_sheet.max_column + 1):
            for row in range(1, values_sheet.max_row + 1):
                value = values_sheet.cell(row, col).value
                if value and " — " in str(value):
                    labels[str(value).split(" — ")[0].strip()] = str(value)

        for row in range(7, intake.max_row + 1):
            field_id = intake.cell(row, 1).value
            if field_id is None:
                continue
            field_id = str(field_id).strip()
            if field_id in flat_answers:
                answer = flat_answers[field_id]
                intake.cell(row, 5).value = labels.get(str(answer), answer)

        wb.save(xlsx)
        _normalize_xlsx_core_properties(xlsx)


def _regenerate_happy_path_exemplar(tmp_path: Path) -> Path:
    workspace = tmp_path / "sample_object_01"
    workspace.mkdir()
    shutil.copy(HAPPY_PATH / "role_assignments.yaml", workspace / "role_assignments.yaml")

    generate(
        workspace,
        project_root=PROJECT_ROOT,
        generated_on=FIXED_DATE.date(),
    )
    _fill_generated_workbooks_from_questionnaire(workspace)
    result = compile_intake(
        workspace,
        project_root=PROJECT_ROOT,
        compiled_on=FIXED_DATE.date(),
    )
    assert result["warnings"] == []
    return workspace


def test_happy_path_golden_file_set_is_explicit_and_complete():
    assert _collect_tracked_happy_path_golden_files() == tuple(sorted(HAPPY_PATH_GOLDEN_FILES))


def test_happy_path_regeneration_matches_checked_in_golden(tmp_path):
    regenerated = _regenerate_happy_path_exemplar(tmp_path)

    for rel_path in HAPPY_PATH_GOLDEN_FILES:
        rel = Path(rel_path)
        expected = HAPPY_PATH / rel
        actual = regenerated / rel
        assert actual.exists(), rel_path
        expected_hash = hashlib.sha256(expected.read_bytes()).hexdigest()
        actual_hash = hashlib.sha256(actual.read_bytes()).hexdigest()
        assert actual_hash == expected_hash, rel_path
