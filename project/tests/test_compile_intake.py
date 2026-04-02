"""Tests for intake compilation (compile_intake.py).

Covers: Excel parsing, status auto-derivation, value validation,
multi-person section assembly, conflicts, normalization,
version/known_unknowns emit, derived .response.yaml, intake_status.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

import pytest
import yaml
from openpyxl import Workbook
from openpyxl.styles import numbers

from intake.compile_intake import (
    _derive_status,
    _parse_value,
    _validate_values,
    _check_conflicts,
    _build_questionnaire,
    _count_statuses,
    compile_intake,
    parse_xlsx,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
FIXED_DATE = date(2026, 4, 2)


# ---------------------------------------------------------------------------
# Helpers for creating test xlsx
# ---------------------------------------------------------------------------

def _make_xlsx(
    path: Path,
    rows: list[tuple[str, str | None, str | None, str | None, str | None]],
) -> None:
    """Create a minimal intake xlsx with given data rows.

    Each row: (field_id, value, status, comment, source_ref).
    Rows 1-5 are filler, row 6 is header, data starts at row 7.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "intake"

    # Title rows (1-5)
    ws["A1"] = "Intake Sheet — Test"
    ws["A2"] = "Roles: test"
    ws["A3"] = "Instructions"
    ws["A4"] = "Guide path"
    # row 5 empty

    # Header row 6
    for col, label in enumerate(
        ["Field ID", "Q", "S", "Ph", "Value", "Status", "Comment", "Source"], 1
    ):
        c = ws.cell(row=6, column=col, value=label)
        c.number_format = numbers.FORMAT_TEXT

    # Data rows
    for i, (fid, val, status, comment, source) in enumerate(rows, 7):
        ws.cell(row=i, column=1, value=fid).number_format = numbers.FORMAT_TEXT
        ws.cell(row=i, column=2, value=fid).number_format = numbers.FORMAT_TEXT
        ws.cell(row=i, column=3, value="S3").number_format = numbers.FORMAT_TEXT
        ws.cell(row=i, column=4, value=2).number_format = numbers.FORMAT_TEXT
        if val is not None:
            ws.cell(row=i, column=5, value=val).number_format = numbers.FORMAT_TEXT
        if status is not None:
            ws.cell(row=i, column=6, value=status).number_format = numbers.FORMAT_TEXT
        if comment is not None:
            ws.cell(row=i, column=7, value=comment).number_format = numbers.FORMAT_TEXT
        if source is not None:
            ws.cell(row=i, column=8, value=source).number_format = numbers.FORMAT_TEXT

    # Minimal _values sheet (required by workbook structure)
    wb.create_sheet("_values")
    wb.create_sheet("_reference")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _make_workspace(
    tmp_path: Path,
    person_rows: dict[str, list[tuple]],
) -> Path:
    """Create a temp workspace with role_assignments and filled xlsx files."""
    import shutil

    ws = tmp_path / "ws"
    ws.mkdir()
    shutil.copy(
        PROJECT_ROOT / "examples" / "sample_object_01" / "role_assignments.yaml",
        ws / "role_assignments.yaml",
    )

    for person_id, rows in person_rows.items():
        _make_xlsx(ws / "intake" / "responses" / f"{person_id}.xlsx", rows)

    return ws


# ---------------------------------------------------------------------------
# Unit tests: _parse_value
# ---------------------------------------------------------------------------


class TestParseValue:
    def test_descriptive_label(self):
        assert _parse_value("segmented \u2014 \u0421\u0435\u0433\u043c\u0435\u043d\u0442\u043e\u0432\u0430\u043d\u0430") == "segmented"

    def test_code_only(self):
        assert _parse_value("segmented") == "segmented"

    def test_none(self):
        assert _parse_value(None) is None

    def test_empty_string(self):
        assert _parse_value("") is None

    def test_whitespace(self):
        assert _parse_value("  ") is None

    def test_label_with_comma(self):
        raw = "hybrid_controlled \u2014 \u0427\u0430\u0441\u0442\u043a\u043e\u0432\u043e \u0441\u043f\u0456\u043b\u044c\u043d\u0438\u0439, \u0447\u0430\u0441\u0442\u043a\u043e\u0432\u043e \u0440\u043e\u0437\u0434\u0456\u043b\u0435\u043d\u0438\u0439"
        assert _parse_value(raw) == "hybrid_controlled"

    def test_integer_string(self):
        assert _parse_value("36") == "36"

    def test_free_text(self):
        assert _parse_value("Sample Industrial Site") == "Sample Industrial Site"


# ---------------------------------------------------------------------------
# Unit tests: _derive_status
# ---------------------------------------------------------------------------


class TestDeriveStatus:
    def test_value_filled_status_empty(self):
        assert _derive_status("yes", None) == ("answered", False)

    def test_value_empty_status_tbd(self):
        assert _derive_status(None, "tbd") == ("tbd", False)

    def test_value_empty_status_not_applicable(self):
        assert _derive_status(None, "not_applicable") == ("not_applicable", False)

    def test_value_empty_status_empty(self):
        assert _derive_status(None, None) == ("unanswered", False)

    def test_value_filled_status_tbd_warns(self):
        status, warned = _derive_status("yes", "tbd")
        assert status == "answered"
        assert warned is True

    def test_blank_string_value(self):
        assert _derive_status("", None) == ("unanswered", False)

    def test_whitespace_status(self):
        assert _derive_status("yes", "  ") == ("answered", False)


# ---------------------------------------------------------------------------
# Unit tests: parse_xlsx
# ---------------------------------------------------------------------------


class TestParseXlsx:
    def test_basic_parsing(self, tmp_path):
        xlsx = tmp_path / "test_person.xlsx"
        _make_xlsx(xlsx, [
            ("wan_required", "yes \u2014 \u0422\u0430\u043a", None, "note", "doc1"),
            ("object_id", "station_01", None, None, None),
        ])
        pid, fields, warnings = parse_xlsx(xlsx)
        assert pid == "test_person"
        assert fields["wan_required"]["value"] == "yes"
        assert fields["wan_required"]["status"] == "answered"
        assert fields["wan_required"]["comment"] == "note"
        assert fields["wan_required"]["source_ref"] == "doc1"
        assert fields["object_id"]["value"] == "station_01"

    def test_empty_row_skipped(self, tmp_path):
        xlsx = tmp_path / "p.xlsx"
        _make_xlsx(xlsx, [("wan_required", "yes", None, None, None)])
        _, fields, _ = parse_xlsx(xlsx)
        assert len(fields) == 1

    def test_warning_on_value_plus_tbd(self, tmp_path):
        xlsx = tmp_path / "p.xlsx"
        _make_xlsx(xlsx, [("wan_required", "yes", "tbd", None, None)])
        _, _, warnings = parse_xlsx(xlsx)
        assert len(warnings) == 1
        assert "tbd" in warnings[0]


# ---------------------------------------------------------------------------
# Unit tests: validation
# ---------------------------------------------------------------------------


class TestValidation:
    def test_valid_enum_passes(self):
        all_fields = {
            "object_type": {"value": "generation", "status": "answered"},
        }
        field_index = {
            "object_type": {"type": "enum", "allowed_values_ref": "object_type"},
        }
        value_dicts = {
            "object_type": [
                {"value_code": "generation"},
                {"value_code": "substation"},
            ],
        }
        errors = _validate_values(all_fields, field_index, value_dicts)
        assert errors == []

    def test_invalid_enum_fails(self):
        all_fields = {
            "object_type": {"value": "spaceship", "status": "answered"},
        }
        field_index = {
            "object_type": {"type": "enum", "allowed_values_ref": "object_type"},
        }
        value_dicts = {
            "object_type": [{"value_code": "generation"}],
        }
        errors = _validate_values(all_fields, field_index, value_dicts)
        assert len(errors) == 1
        assert "spaceship" in errors[0]

    def test_tbd_always_valid(self):
        all_fields = {
            "object_type": {"value": "tbd", "status": "answered"},
        }
        field_index = {
            "object_type": {"type": "enum", "allowed_values_ref": "object_type"},
        }
        value_dicts = {
            "object_type": [{"value_code": "generation"}],
        }
        errors = _validate_values(all_fields, field_index, value_dicts)
        assert errors == []

    def test_integer_valid(self):
        all_fields = {
            "growth_horizon_months": {"value": "36", "status": "answered"},
        }
        field_index = {"growth_horizon_months": {"type": "integer"}}
        errors = _validate_values(all_fields, field_index, {})
        assert errors == []

    def test_integer_invalid(self):
        all_fields = {
            "growth_horizon_months": {"value": "abc", "status": "answered"},
        }
        field_index = {"growth_horizon_months": {"type": "integer"}}
        errors = _validate_values(all_fields, field_index, {})
        assert len(errors) == 1

    def test_string_skipped(self):
        all_fields = {
            "object_id": {"value": "anything goes", "status": "answered"},
        }
        field_index = {"object_id": {"type": "string"}}
        errors = _validate_values(all_fields, field_index, {})
        assert errors == []

    def test_unanswered_skipped(self):
        all_fields = {
            "object_type": {"value": None, "status": "unanswered"},
        }
        field_index = {
            "object_type": {"type": "enum", "allowed_values_ref": "object_type"},
        }
        errors = _validate_values(all_fields, field_index, {})
        assert errors == []


# ---------------------------------------------------------------------------
# Unit tests: conflicts
# ---------------------------------------------------------------------------


class TestConflicts:
    def test_no_conflict(self):
        person_fields = {
            "a": {"f1": {"status": "answered"}},
            "b": {"f2": {"status": "answered"}},
        }
        assert _check_conflicts(person_fields) == []

    def test_conflict_detected(self):
        person_fields = {
            "a": {"f1": {"status": "answered"}},
            "b": {"f1": {"status": "answered"}},
        }
        errors = _check_conflicts(person_fields)
        assert len(errors) == 1
        assert "f1" in errors[0]

    def test_unanswered_no_conflict(self):
        person_fields = {
            "a": {"f1": {"status": "unanswered"}},
            "b": {"f1": {"status": "answered"}},
        }
        assert _check_conflicts(person_fields) == []


# ---------------------------------------------------------------------------
# Unit tests: questionnaire builder
# ---------------------------------------------------------------------------


class TestBuildQuestionnaire:
    def test_version_and_known_unknowns(self):
        core_data = {"sections": [{"id": "metadata", "fields": []}, {"id": "known_unknowns", "fields": []}]}
        q = _build_questionnaire({}, {}, core_data)
        assert q["version"] == "0.2.0"
        assert q["known_unknowns"] == {}

    def test_answered_field_emitted(self):
        all_fields = {"object_type": {"value": "generation", "status": "answered"}}
        field_index = {"object_type": {"type": "enum"}}
        core_data = {"sections": [{"id": "metadata", "fields": ["object_type"]}, {"id": "known_unknowns"}]}
        q = _build_questionnaire(all_fields, field_index, core_data)
        assert q["metadata"]["object_type"] == "generation"

    def test_unanswered_field_omitted(self):
        all_fields = {"object_type": {"value": None, "status": "unanswered"}}
        field_index = {"object_type": {"type": "enum"}}
        core_data = {"sections": [{"id": "metadata", "fields": ["object_type"]}, {"id": "known_unknowns"}]}
        q = _build_questionnaire(all_fields, field_index, core_data)
        assert "metadata" not in q  # empty section omitted

    def test_tbd_emitted_as_string(self):
        all_fields = {"object_type": {"value": None, "status": "tbd"}}
        field_index = {"object_type": {"type": "enum"}}
        core_data = {"sections": [{"id": "metadata", "fields": ["object_type"]}, {"id": "known_unknowns"}]}
        q = _build_questionnaire(all_fields, field_index, core_data)
        assert q["metadata"]["object_type"] == "tbd"

    def test_not_applicable_emitted_as_tbd(self):
        all_fields = {"object_type": {"value": None, "status": "not_applicable"}}
        field_index = {"object_type": {"type": "enum"}}
        core_data = {"sections": [{"id": "metadata", "fields": ["object_type"]}, {"id": "known_unknowns"}]}
        q = _build_questionnaire(all_fields, field_index, core_data)
        assert q["metadata"]["object_type"] == "tbd"

    def test_integer_field_emitted_as_int(self):
        all_fields = {"growth_horizon_months": {"value": "36", "status": "answered"}}
        field_index = {"growth_horizon_months": {"type": "integer"}}
        core_data = {"sections": [{"id": "object_profile", "fields": ["growth_horizon_months"]}, {"id": "known_unknowns"}]}
        q = _build_questionnaire(all_fields, field_index, core_data)
        assert q["object_profile"]["growth_horizon_months"] == 36
        assert isinstance(q["object_profile"]["growth_horizon_months"], int)

    def test_section_order_preserved(self):
        all_fields = {
            "wan_required": {"value": "yes", "status": "answered"},
            "object_type": {"value": "generation", "status": "answered"},
        }
        field_index = {
            "wan_required": {"type": "enum"},
            "object_type": {"type": "enum"},
        }
        core_data = {"sections": [
            {"id": "metadata", "fields": ["object_type"]},
            {"id": "external_transport", "fields": ["wan_required"]},
            {"id": "known_unknowns"},
        ]}
        q = _build_questionnaire(all_fields, field_index, core_data)
        keys = list(q.keys())
        assert keys.index("metadata") < keys.index("external_transport")
        assert keys[-1] == "known_unknowns"


# ---------------------------------------------------------------------------
# Integration: full compile
# ---------------------------------------------------------------------------


class TestCompileIntegration:
    """Full generate → fill → compile cycle using sample_object_01."""

    @pytest.fixture(scope="class")
    def compiled(self, tmp_path_factory):
        import shutil
        from intake.generate_intake_sheets import generate

        tmp = tmp_path_factory.mktemp("compile")
        ws = tmp / "sample_object_01"
        ws.mkdir()
        shutil.copy(
            PROJECT_ROOT / "examples" / "sample_object_01" / "role_assignments.yaml",
            ws / "role_assignments.yaml",
        )

        # Generate
        generate(ws, project_root=PROJECT_ROOT, generated_on=FIXED_DATE)

        # Fill with existing questionnaire values
        from openpyxl import load_workbook as lw
        existing = yaml.safe_load(
            (PROJECT_ROOT / "examples" / "sample_object_01" / "questionnaire.yaml")
            .read_text(encoding="utf-8")
        )
        flat: dict[str, Any] = {}
        for k, v in existing.items():
            if isinstance(v, dict) and k not in ("version", "known_unknowns"):
                flat.update(v)

        for xlsx in sorted((ws / "intake" / "responses").glob("*.xlsx")):
            wb = lw(xlsx)
            wsi, wsv = wb["intake"], wb["_values"]
            labels: dict[str, str] = {}
            for c in range(1, 30):
                for r in range(1, 20):
                    v = wsv.cell(r, c).value
                    if v and " \u2014 " in str(v):
                        labels[str(v).split(" \u2014 ")[0].strip()] = v
            for r in range(7, wsi.max_row + 1):
                fid = wsi.cell(r, 1).value
                if not fid:
                    break
                fid = str(fid).strip()
                if fid in flat:
                    val = flat[fid]
                    wsi.cell(r, 5).value = labels.get(str(val), str(val))
            wb.save(xlsx)

        # Compile
        return (
            compile_intake(ws, project_root=PROJECT_ROOT, compiled_on=FIXED_DATE),
            ws,
            existing,
        )

    def test_questionnaire_version(self, compiled):
        result, _, _ = compiled
        assert result["questionnaire"]["version"] == "0.2.0"

    def test_questionnaire_known_unknowns(self, compiled):
        result, _, _ = compiled
        assert result["questionnaire"]["known_unknowns"] == {}

    def test_all_41_fields_accounted(self, compiled):
        result, _, _ = compiled
        assert len(result["all_fields"]) == 41

    def test_all_sections_present(self, compiled):
        result, _, _ = compiled
        q = result["questionnaire"]
        expected_sections = [
            "metadata", "object_profile", "critical_services",
            "external_transport", "security_access", "time_sync",
            "power_environment", "resilience", "operations",
            "acceptance_criteria", "governance",
        ]
        for s in expected_sections:
            assert s in q, f"Section {s} missing"

    def test_values_match_existing_questionnaire(self, compiled):
        result, _, existing = compiled
        q = result["questionnaire"]
        for section_key, section_data in existing.items():
            if not isinstance(section_data, dict):
                continue
            if section_key == "known_unknowns":
                continue
            for fid, expected_val in section_data.items():
                actual = q.get(section_key, {}).get(fid)
                assert str(actual) == str(expected_val), (
                    f"{section_key}.{fid}: expected={expected_val} got={actual}"
                )

    def test_questionnaire_yaml_written(self, compiled):
        _, ws, _ = compiled
        assert (ws / "questionnaire.yaml").exists()

    def test_intake_status_yaml_written(self, compiled):
        _, ws, _ = compiled
        path = ws / "reports" / "intake_status.yaml"
        assert path.exists()
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
        assert data["object_id"] == "sample_object_01"
        assert data["totals"]["total"] == 41
        assert data["totals"]["answered"] == 41

    def test_intake_status_md_written(self, compiled):
        _, ws, _ = compiled
        path = ws / "reports" / "intake_status.md"
        assert path.exists()
        text = path.read_text(encoding="utf-8")
        assert "Intake Status" in text
        assert f"Compiled at: {FIXED_DATE.isoformat()}" in text
        assert "41/41" in text
        assert "## Per Person" in text
        assert "## Phase Readiness" in text
        assert "## Field Ownership Table" in text

    def test_response_yaml_per_person(self, compiled):
        _, ws, _ = compiled
        responses = ws / "intake" / "responses"
        for xlsx in responses.glob("*.xlsx"):
            pid = xlsx.stem
            resp = responses / f"{pid}.response.yaml"
            assert resp.exists(), f"Missing response yaml for {pid}"
            data = yaml.safe_load(resp.read_text(encoding="utf-8"))
            assert data["person_id"] == pid
            assert "fields" in data

    def test_fixed_compile_date_propagated_to_outputs(self, compiled):
        _, ws, _ = compiled
        status = yaml.safe_load(
            (ws / "reports" / "intake_status.yaml").read_text(encoding="utf-8")
        )
        assert status["compiled_at"] == FIXED_DATE.isoformat()

        response = yaml.safe_load(
            (
                ws
                / "intake"
                / "responses"
                / "sample_arch.response.yaml"
            ).read_text(encoding="utf-8")
        )
        assert response["compiled_at"] == FIXED_DATE.isoformat()

    def test_no_warnings(self, compiled):
        result, _, _ = compiled
        assert result["warnings"] == []

    def test_scope_summary_in_status_md(self, compiled):
        _, ws, _ = compiled
        text = (ws / "reports" / "intake_status.md").read_text(encoding="utf-8")
        assert "generation" in text
        assert "high" in text

    def test_phase_readiness_complete(self, compiled):
        _, ws, _ = compiled
        text = (ws / "reports" / "intake_status.md").read_text(encoding="utf-8")
        assert "Phase 1 (Identity): complete" in text
        assert "Phase 2 (Constraints): complete" in text
        assert "Phase 3 (Operations): complete" in text


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    def test_compile_raises_on_invalid_value(self, tmp_path):
        ws = _make_workspace(tmp_path, {
            "sample_arch": [
                ("wan_required", "INVALID_VALUE", None, None, None),
            ],
        })
        with pytest.raises(ValueError, match="not in"):
            compile_intake(ws, project_root=PROJECT_ROOT)

    def test_compile_raises_on_conflict(self, tmp_path):
        ws = _make_workspace(tmp_path, {
            "person_a": [
                ("wan_required", "yes", None, None, None),
            ],
            "person_b": [
                ("wan_required", "no", None, None, None),
            ],
        })
        with pytest.raises(ValueError, match="multiple persons"):
            compile_intake(ws, project_root=PROJECT_ROOT)

    def test_partial_fill_produces_tbd_and_unanswered(self, tmp_path):
        ws = _make_workspace(tmp_path, {
            "sample_pm_owner": [
                ("object_id", "test_01", None, None, None),
                ("object_type", None, "tbd", None, None),
            ],
        })
        result = compile_intake(ws, project_root=PROJECT_ROOT)
        af = result["all_fields"]
        assert af["object_id"]["status"] == "answered"
        assert af["object_type"]["status"] == "tbd"

        q = result["questionnaire"]
        assert q["metadata"]["object_id"] == "test_01"
        assert q["metadata"]["object_type"] == "tbd"

    def test_not_applicable_emits_tbd_in_questionnaire(self, tmp_path):
        ws = _make_workspace(tmp_path, {
            "sample_arch": [
                ("wan_required", None, "not_applicable", None, None),
            ],
        })
        result = compile_intake(ws, project_root=PROJECT_ROOT)
        af = result["all_fields"]
        assert af["wan_required"]["status"] == "not_applicable"

        q = result["questionnaire"]
        assert q["external_transport"]["wan_required"] == "tbd"

    def test_no_xlsx_raises(self, tmp_path):
        import shutil
        ws = tmp_path / "empty_ws"
        ws.mkdir()
        shutil.copy(
            PROJECT_ROOT / "examples" / "sample_object_01" / "role_assignments.yaml",
            ws / "role_assignments.yaml",
        )
        (ws / "intake" / "responses").mkdir(parents=True)
        with pytest.raises(FileNotFoundError):
            compile_intake(ws, project_root=PROJECT_ROOT)
